import base64
import hashlib
import http.client
import importlib.util
import json
import os
import subprocess
import sys
import tempfile
import threading
import time
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook
from PIL import Image
import engine_config as lifecycle_engine_config
from profile_runtime import (
    ProfileRuntimeError,
    _harden_windows_secret_acl,
    atomic_write_json,
    create_profile_directories,
    resolve_profile_runtime,
    write_private_single_line,
)


ROOT = Path(__file__).resolve().parents[1]
SPEC = importlib.util.spec_from_file_location(
    "runtime_security_service", ROOT / "service" / "ocr_service.py"
)
service = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(service)


class RuntimeHttpSecurityTest(unittest.TestCase):
    TOKEN = "T" * 64
    OWNER = "a" * 64
    INSTANCE = "b" * 32

    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        root = Path(self.temp.name)
        self.inbox = root / "inbox"
        self.jobs = root / "jobs"
        self.inbox.mkdir()
        self.jobs.mkdir()
        self.old_jobs = service.JOBS_DIR
        service.JOBS_DIR = self.jobs
        self.server = service.build_server(
            "127.0.0.1",
            0,
            auth_token=self.TOKEN,
            profile="default",
            data_root=root,
            owner_fingerprint=self.OWNER,
            instance_id=self.INSTANCE,
            started_at="2026-08-10T00:00:00Z",
            runtime_identity=None,
        )
        self.server.inbox = self.inbox
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)
        self.thread.start()

    def tearDown(self):
        self.server.shutdown()
        self.server.server_close()
        self.thread.join(timeout=2)
        service.JOBS_DIR = self.old_jobs
        self.temp.cleanup()

    def request(self, method, path, body=None, headers=None):
        connection = http.client.HTTPConnection(
            "127.0.0.1", self.server.server_port, timeout=3
        )
        connection.request(method, path, body=body, headers=headers or {})
        response = connection.getresponse()
        payload = response.read()
        result = response.status, dict(response.getheaders()), payload
        connection.close()
        return result

    def token_headers(self, *, include_owner=True, **extra):
        value = {"Authorization": f"Bearer {self.TOKEN}"}
        if include_owner:
            value["X-Document-Reader-Owner"] = self.OWNER
        value.update(extra)
        return value

    def test_health_requires_token_and_exact_owner_then_attests_exact_schema(self):
        status, _, _ = self.request("GET", "/api/health")
        self.assertEqual(status, 401)
        status, _, _ = self.request(
            "GET", "/api/health", headers=self.token_headers(include_owner=False)
        )
        self.assertEqual(status, 403)
        status, headers, body = self.request(
            "GET",
            "/api/health",
            headers=self.token_headers(**{"X-Document-Reader-Owner": self.OWNER}),
        )
        self.assertEqual(status, 200)
        health = json.loads(body)
        self.assertEqual(
            set(health),
            {
                "status", "service", "version", "api_version", "profile_name",
                "owner_fingerprint", "instance_id", "port", "pid", "started_at",
            },
        )
        self.assertEqual(health["service"], "hermes-document-reader")
        self.assertEqual(health["version"], "0.1.0")
        self.assertEqual(health["profile_name"], "default")
        self.assertEqual(health["owner_fingerprint"], self.OWNER)
        self.assertEqual(health["port"], self.server.server_port)
        self.assertIn("private, no-store", headers["Cache-Control"])

    def test_basic_bearer_and_x_header_authentication(self):
        basic = base64.b64encode(f"default:{self.TOKEN}".encode()).decode()
        for headers in (
            {
                "Authorization": f"Basic {basic}",
                "X-Document-Reader-Owner": self.OWNER,
            },
            {
                "Authorization": f"Bearer {self.TOKEN}",
                "X-Document-Reader-Owner": self.OWNER,
            },
            {
                "X-Document-Reader-Token": self.TOKEN,
                "X-Document-Reader-Owner": self.OWNER,
            },
        ):
            status, _, _ = self.request("GET", "/api/state", headers=headers)
            self.assertEqual(status, 200)

    def test_duplicate_authentication_headers_fail_closed(self):
        connection = http.client.HTTPConnection(
            "127.0.0.1", self.server.server_port, timeout=3
        )
        connection.putrequest("GET", "/api/state", skip_host=True)
        connection.putheader("Host", f"127.0.0.1:{self.server.server_port}")
        connection.putheader("Authorization", f"Bearer {self.TOKEN}")
        connection.putheader("Authorization", f"Bearer {self.TOKEN}")
        connection.putheader("X-Document-Reader-Owner", self.OWNER)
        connection.endheaders()
        response = connection.getresponse()
        response.read()
        self.assertEqual(response.status, 401)
        connection.close()

    def test_host_and_origin_are_fail_closed(self):
        status, _, _ = self.request(
            "GET", "/api/state", headers=self.token_headers(Host="evil.example")
        )
        self.assertEqual(status, 421)
        status, _, _ = self.request(
            "GET", "/api/state", headers=self.token_headers(Origin="null")
        )
        self.assertEqual(status, 403)

    def test_reused_token_with_wrong_profile_owner_cannot_read_state_or_jobs(self):
        wrong = self.token_headers(**{"X-Document-Reader-Owner": "c" * 64})
        status, _, _ = self.request("GET", "/api/state", headers=wrong)
        self.assertEqual(status, 403)
        job_id = "20260810-010101-aaaaaaaa"
        job = self.jobs / job_id
        job.mkdir()
        (job / "out.txt").write_text("private", encoding="utf-8")
        status, _, _ = self.request(
            "GET", f"/jobs/{job_id}/out.txt", headers=wrong
        )
        self.assertEqual(status, 403)

    def test_encoded_routes_share_one_canonical_authorization_boundary(self):
        token_only = self.token_headers(include_owner=False)
        for path in ("/%61pi/state", "/%61pi/health", "/%6aobs/20260810-010101-aaaaaaaa/out.txt"):
            status, _, _ = self.request("GET", path, headers=token_only)
            self.assertEqual(status, 403, path)

        status, _, body = self.request(
            "GET", "/%61pi/state", headers=self.token_headers()
        )
        self.assertEqual(status, 200)
        self.assertEqual(json.loads(body)["profile"], "default")
        status, _, body = self.request(
            "GET", "/%61pi/health", headers=self.token_headers()
        )
        self.assertEqual(status, 200)
        self.assertEqual(json.loads(body)["owner_fingerprint"], self.OWNER)

        status, _, _ = self.request(
            "POST", "/%61pi/cancel", headers=token_only
        )
        self.assertEqual(status, 403)
        status, _, _ = self.request(
            "POST", "/%61pi/cancel", headers=self.token_headers()
        )
        self.assertEqual(status, 409)

    def test_request_target_rejects_invalid_escapes_and_never_double_decodes(self):
        for path in ("/%", "/%6gpi/state", "/%ffapi/state"):
            status, _, _ = self.request("GET", path, headers=self.token_headers())
            self.assertEqual(status, 400, path)
        for path in ("/%2561pi/state", "/%252561pi/state"):
            status, _, _ = self.request("GET", path, headers=self.token_headers())
            self.assertEqual(status, 404, path)

    def test_browser_owner_cookie_is_scoped_to_the_expected_instance(self):
        status, headers, _ = self.request("GET", "/", headers=self.token_headers())
        self.assertEqual(status, 200)
        cookie = headers["Set-Cookie"].split(";", 1)[0]
        status, _, _ = self.request(
            "GET",
            "/api/state",
            headers={"Authorization": f"Bearer {self.TOKEN}", "Cookie": cookie},
        )
        self.assertEqual(status, 200)
        status, _, _ = self.request(
            "GET",
            "/api/health",
            headers={"Authorization": f"Bearer {self.TOKEN}", "Cookie": cookie},
        )
        self.assertEqual(status, 403)
        status, _, _ = self.request(
            "GET",
            "/api/state",
            headers=self.token_headers(Origin="http://127.0.0.1:1"),
        )
        self.assertEqual(status, 403)

    def test_upload_rejects_mismatched_content_and_leaves_no_residue(self):
        tracked = {}
        for index in range(32):
            name = f"malicious-{index}.pdf"
            status, _, _ = self.request(
                "POST",
                f"/api/upload?name={name}",
                body=b"not a pdf",
                headers=self.token_headers(**{"Content-Length": "9"}),
            )
            self.assertEqual(status, 415)
            tracked[self.inbox / name] = 0
        self.assertEqual(list(self.inbox.iterdir()), [])
        service.prune_watcher_sizes(tracked, set(self.inbox.iterdir()))
        self.assertEqual(tracked, {})

    def test_jobs_require_authentication_and_never_public_cache(self):
        job_id = "20260810-010101-aaaaaaaa"
        job = self.jobs / job_id
        job.mkdir()
        (job / "out.txt").write_text("private", encoding="utf-8")
        status, _, _ = self.request("GET", f"/jobs/{job_id}/out.txt")
        self.assertEqual(status, 401)
        status, headers, body = self.request(
            "GET", f"/jobs/{job_id}/out.txt", headers=self.token_headers()
        )
        self.assertEqual(status, 200)
        self.assertEqual(body, b"private")
        self.assertIn("private, no-store", headers["Cache-Control"])

    def test_job_download_streams_the_opened_identity_after_path_replacement(self):
        job_id = "20260810-010101-aaaaaaaa"
        job = self.jobs / job_id
        job.mkdir()
        target = job / "out.txt"
        target.write_bytes(b"approved")
        original_open = service._open_regular_readonly

        def replace_after_open(path, expected=None):
            descriptor, info = original_open(path, expected)
            if Path(path) == target:
                target.unlink()
                target.write_bytes(b"replacement")
            return descriptor, info

        with mock.patch.object(
            service, "_open_regular_readonly", side_effect=replace_after_open
        ):
            status, _, body = self.request(
                "GET", f"/jobs/{job_id}/out.txt", headers=self.token_headers()
            )
        self.assertEqual(status, 200)
        self.assertEqual(body, b"approved")

    def test_upload_has_no_visible_final_reservation_and_directory_collision_is_safe(self):
        injected = self.inbox / "scan.pdf"
        injected.mkdir()
        (injected / "sentinel").write_text("keep", encoding="utf-8")
        body = b"%PDF-test"
        status, _, payload = self.request(
            "POST",
            "/api/upload?name=scan.pdf",
            body=body,
            headers=self.token_headers(**{"Content-Length": str(len(body))}),
        )
        self.assertEqual(status, 200, payload)
        self.assertEqual(json.loads(payload)["name"], "scan (1).pdf")
        self.assertTrue((injected / "sentinel").is_file())
        self.assertEqual((self.inbox / "scan (1).pdf").read_bytes(), body)
        self.assertEqual(list(self.inbox.glob("*.uploading")), [])


class RuntimeFileSafetyTest(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.saved = {
            name: getattr(service, name)
            for name in (
                "PROFILE_ID", "DATA_ROOT", "STATE_DIR", "JOBS_DIR", "PROCESSED_DIR",
                "ON_HOLD_DIR", "NEEDS_REVIEW_DIR", "QUARANTINE_DIR", "HISTORY_PATH",
                "_RUNTIME_IDENTITY", "_RUNTIME_CHILD_IDENTITIES",
            )
        }
        self.saved_history = service.STATE["history"]
        self.saved_job = service.STATE["job"]
        self.saved_queue = service.STATE["queue"]
        service.STATE["history"] = []
        service.STATE["job"] = None
        service.STATE["queue"] = []
        service._retry_after.clear()
        service._retry_counts.clear()
        service.configure_runtime("default", self.root / "data")

    def history_entry(self, job_id="20260810-010101-aaaaaaaa"):
        return {
            "id": job_id,
            "name": "private.pdf",
            "status": "finished",
            "when": "2026-08-10 01:01",
            "pages": 1,
            "errors": 0,
            "secs": 1.0,
            "chars": 12,
            "links": {
                "md": f"/jobs/{job_id}/private.txt",
                "xlsx": f"/jobs/{job_id}/private.xlsx",
            },
        }

    def tearDown(self):
        for name, value in self.saved.items():
            setattr(service, name, value)
        service.STATE["history"] = self.saved_history
        service.STATE["job"] = self.saved_job
        service.STATE["queue"] = self.saved_queue
        service._retry_after.clear()
        service._retry_counts.clear()
        self.temp.cleanup()

    def test_atomic_history_has_no_temporary_residue(self):
        service.STATE["history"] = [self.history_entry()]
        service.save_history()
        self.assertEqual(
            json.loads(service.HISTORY_PATH.read_text())[0]["id"],
            "20260810-010101-aaaaaaaa",
        )
        self.assertEqual(list(service.STATE_DIR.glob("*.tmp")), [])

    def test_corrupt_history_is_quarantined_and_never_silently_overwritten(self):
        service.HISTORY_PATH.write_text("{not-json", encoding="utf-8")
        service.load_history()
        quarantined = list(service.STATE_DIR.glob("history.corrupt-*.json"))
        self.assertEqual(len(quarantined), 1)
        self.assertEqual(quarantined[0].read_text(encoding="utf-8"), "{not-json")
        self.assertFalse(service.HISTORY_PATH.exists())
        service.STATE["history"] = [
            self.history_entry("20260810-010102-bbbbbbbb")
        ]
        service.save_history()
        self.assertTrue(quarantined[0].exists())

    def test_corrupt_history_refuses_reset_when_quarantine_fails(self):
        service.HISTORY_PATH.write_text("{not-json", encoding="utf-8")
        service.STATE["history"] = [{"id": "preserve-in-memory"}]
        with mock.patch.object(service.os, "replace", side_effect=OSError("locked")):
            with self.assertRaisesRegex(RuntimeError, "refusing to overwrite"):
                service.load_history()
        self.assertEqual(service.STATE["history"], [{"id": "preserve-in-memory"}])
        self.assertTrue(service.HISTORY_PATH.exists())

    def test_history_migration_strips_absolute_paths_before_state_or_storage(self):
        entry = self.history_entry()
        entry["name"] = "private-folder/private.pdf"
        entry["paths"] = {
            "folder": "C:/Users/private/document-reader/data/processed",
            "original": "C:/Users/private/document-reader/data/processed/private.pdf",
        }
        service.HISTORY_PATH.write_text(json.dumps([entry]), encoding="utf-8")
        service.load_history()
        self.assertEqual(service.STATE["history"][0]["name"], "private.pdf")
        self.assertNotIn("paths", service.STATE["history"][0])
        service.save_history()
        rendered = service.HISTORY_PATH.read_text(encoding="utf-8")
        self.assertNotIn("C:/Users/private", rendered)

    def test_owner_lock_is_exclusive_and_reusable(self):
        path = self.root / "runtime" / "service.lock"
        first = service.RuntimeOwnerLock(path).acquire()
        try:
            with self.assertRaises(RuntimeError):
                service.RuntimeOwnerLock(path).acquire()
        finally:
            first.release()
        service.RuntimeOwnerLock(path).acquire().release()

    def test_service_token_is_revalidated_as_an_exact_private_value(self):
        token_path = (self.root / "config" / "service.token").resolve()
        token_path.parent.mkdir()
        token = "S" * 64
        write_private_single_line(token_path, token, minimum=43, maximum=128)
        self.assertEqual(service.read_service_token(token_path), token)
        token_path.write_text(token + " \n", encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "exact bounded line"):
            service.read_service_token(token_path)

    def test_owner_lock_remains_exclusive_until_old_process_dies(self):
        lock_path = self.root / "runtime-process" / "service.lock"
        ready = self.root / "runtime-process" / "ready"
        child_code = (
            "import importlib.util,pathlib,sys,time;"
            "spec=importlib.util.spec_from_file_location('lock_service',sys.argv[1]);"
            "m=importlib.util.module_from_spec(spec);spec.loader.exec_module(m);"
            "lock=m.RuntimeOwnerLock(pathlib.Path(sys.argv[2])).acquire();"
            "pathlib.Path(sys.argv[3]).write_text('ready');"
            "time.sleep(60)"
        )
        child = subprocess.Popen(
            [sys.executable, "-c", child_code, str(ROOT / "service" / "ocr_service.py"), str(lock_path), str(ready)]
        )
        try:
            deadline = time.time() + 10
            while not ready.exists() and child.poll() is None and time.time() < deadline:
                time.sleep(0.05)
            self.assertTrue(ready.exists(), "child did not acquire the runtime lock")
            with self.assertRaises(RuntimeError):
                service.RuntimeOwnerLock(lock_path).acquire()
        finally:
            child.terminate()
            child.wait(timeout=10)
        replacement = None
        deadline = time.time() + 5
        while replacement is None and time.time() < deadline:
            try:
                replacement = service.RuntimeOwnerLock(lock_path).acquire()
            except RuntimeError:
                time.sleep(0.05)
        self.assertIsNotNone(replacement, "owner lock survived the old process")
        replacement.release()

    def test_hard_shutdown_kills_delayed_mutator_before_releasing_owner_lock(self):
        runtime = self.root / "shutdown-runtime"
        lock_path = runtime / "service.lock"
        ready = runtime / "ready.json"
        marker = runtime / "late-mutation"
        child_code = """
import importlib.util, json, pathlib, sys, threading, time
spec = importlib.util.spec_from_file_location('shutdown_service', sys.argv[1])
m = importlib.util.module_from_spec(spec)
spec.loader.exec_module(m)
runtime = pathlib.Path(sys.argv[2])
runtime.mkdir(parents=True, exist_ok=True)
with m.RuntimeOwnerLock(runtime / 'service.lock'):
    server = m.build_server(
        '127.0.0.1', 0, auth_token='T' * 64, profile='default',
        data_root=runtime / 'data', owner_fingerprint='a' * 64,
        instance_id='b' * 32, started_at='2026-08-10T00:00:00Z',
        runtime_identity=None, hard_exit_on_shutdown=True,
    )
    def mutate_late():
        time.sleep(1)
        (runtime / 'late-mutation').write_text('unsafe')
    threading.Thread(target=mutate_late, daemon=True).start()
    (runtime / 'ready.json').write_text(json.dumps({'port': server.server_port}))
    server.serve_forever()
"""
        child = subprocess.Popen(
            [sys.executable, "-c", child_code, str(ROOT / "service" / "ocr_service.py"), str(runtime)]
        )
        try:
            deadline = time.time() + 10
            while not ready.exists() and child.poll() is None and time.time() < deadline:
                time.sleep(0.05)
            self.assertTrue(ready.exists(), "child service did not become ready")
            port = json.loads(ready.read_text())["port"]
            connection = http.client.HTTPConnection("127.0.0.1", port, timeout=3)
            connection.request(
                "POST",
                "/api/shutdown",
                headers={
                    "Authorization": "Bearer " + "T" * 64,
                    "X-Document-Reader-Owner": "a" * 64,
                },
            )
            response = connection.getresponse()
            response.read()
            connection.close()
            self.assertEqual(response.status, 200)
            child.wait(timeout=10)
            time.sleep(1.1)
            self.assertFalse(marker.exists())
        finally:
            if child.poll() is None:
                child.terminate()
                child.wait(timeout=10)
        service.RuntimeOwnerLock(lock_path).acquire().release()

    def test_permanent_failure_moves_source_to_quarantine_with_reason(self):
        inbox = self.root / "data" / "inbox"
        inbox.mkdir()
        source = inbox / "bad.pdf"
        source.write_bytes(b"not-pdf")
        hostile = (
            "https://engine.secret.invalid/v1/models token=super-secret "
            "C:/Users/private/document.pdf"
        )
        result = service.handle_failed_source(source, inbox, hostile, permanent=True)
        self.assertEqual(result, "quarantined")
        moved = service.QUARANTINE_DIR / "bad.pdf"
        self.assertTrue(moved.is_file())
        reason = json.loads((service.QUARANTINE_DIR / "bad.pdf.error.json").read_text())
        self.assertEqual(reason["error"], "document processing failed")
        serialized = json.dumps(reason)
        for secret in ("engine.secret.invalid", "super-secret", "C:/Users/private"):
            self.assertNotIn(secret, serialized)

    def test_internal_exception_details_are_classified_without_disclosure(self):
        hostile = RuntimeError(
            "POST https://engine.secret.invalid/v1/models failed; "
            "Bearer super-secret; C:/Users/private/document.pdf"
        )
        rendered = service.classified_error(hostile)
        self.assertEqual(rendered, "document processing failed")
        for secret in ("engine.secret.invalid", "super-secret", "C:/Users/private"):
            self.assertNotIn(secret, rendered)

    def test_source_move_failure_never_claims_completion(self):
        source = self.root / "source.pdf"
        source.write_bytes(b"%PDF-test")
        destination = self.root / "destination"
        destination.mkdir()
        with mock.patch.object(service.os, "link", side_effect=OSError("locked")):
            with self.assertRaises(OSError):
                service.move_source_confirmed(source, destination)
        self.assertTrue(source.exists())

    def test_disposition_uses_atomic_no_clobber_names_for_every_owned_destination(self):
        for index, destination in enumerate(
            (
                service.PROCESSED_DIR,
                service.ON_HOLD_DIR,
                service.NEEDS_REVIEW_DIR,
                service.QUARANTINE_DIR,
            )
        ):
            source_dir = self.root / "data" / f"source-{index}"
            source_dir.mkdir()
            source = source_dir / "scan.pdf"
            source.write_bytes(b"%PDF-safe")
            injected = destination / source.name
            injected.mkdir()
            (injected / "sentinel").write_text("keep", encoding="utf-8")
            moved = service.move_source_confirmed(source, destination)
            self.assertEqual(moved.name, "scan (1).pdf")
            self.assertEqual(moved.read_bytes(), b"%PDF-safe")
            self.assertTrue((injected / "sentinel").is_file())
            self.assertFalse(source.exists())

    def test_replaced_child_root_is_rejected_before_source_read_or_retention_delete(self):
        inbox = self.root / "data" / "inbox"
        inbox.mkdir()
        service._RUNTIME_IDENTITY = service.capture_runtime_identity(
            service.DATA_ROOT, "a" * 64, "default"
        )
        service._RUNTIME_CHILD_IDENTITIES = service.capture_runtime_directories(
            {"inbox": inbox, "jobs": service.JOBS_DIR}
        )
        old_inbox = inbox.with_name("inbox-old")
        inbox.rename(old_inbox)
        inbox.mkdir()
        hostile = inbox / "outside.pdf"
        hostile.write_bytes(b"%PDF-outside")
        with self.assertRaises(RuntimeError):
            service.validate_source_file(hostile, inbox)
        hostile.unlink()
        inbox.rmdir()
        old_inbox.rename(inbox)

        old_jobs = service.JOBS_DIR.with_name("jobs-old")
        service.JOBS_DIR.rename(old_jobs)
        service.JOBS_DIR.mkdir()
        job = service.JOBS_DIR / "20260810-010101-aaaaaaaa"
        job.mkdir()
        sentinel = job / "sentinel.txt"
        sentinel.write_text("keep", encoding="utf-8")
        with self.assertRaises(RuntimeError):
            service.enforce_retention()
        self.assertTrue(sentinel.exists())

    def test_pdf_geometry_is_rejected_before_chandra_render(self):
        source = self.root / "oversized.pdf"
        source.write_bytes(b"%PDF-test")

        class Page:
            def get_width(self):
                return 1

            def get_height(self):
                return 100000

            def close(self):
                pass

        class Document:
            def __len__(self):
                return 1

            def __getitem__(self, _index):
                return Page()

            def close(self):
                pass

        with mock.patch.object(
            service.pdfium, "PdfDocument", return_value=Document()
        ), mock.patch.object(service, "load_file") as loader:
            with self.assertRaisesRegex(ValueError, "pre-render"):
                service.load_page(source, 0)
            loader.assert_not_called()

    def test_image_geometry_is_rejected_before_convert_or_resize(self):
        source = self.root / "oversized.png"
        source.write_bytes(b"image")
        image = mock.MagicMock()
        image.__enter__.return_value = image
        image.size = (1, 100000)
        with mock.patch("PIL.Image.open", return_value=image), mock.patch.object(
            service, "load_file"
        ) as loader:
            with self.assertRaisesRegex(ValueError, "pre-render"):
                service.load_page(source, 0)
            loader.assert_not_called()

    def test_ocr_uses_private_snapshot_and_refuses_a_changed_live_source(self):
        inbox = self.root / "data" / "inbox"
        inbox.mkdir(exist_ok=True)
        source = inbox / "scan.png"
        original = b"\x89PNG\r\n\x1a\n" + b"original-payload"
        changed = b"\x89PNG\r\n\x1a\n" + b"modified-payload"
        self.assertEqual(len(original), len(changed))
        source.write_bytes(original)
        loaded_paths = []

        loaded_bytes = []

        def load_snapshot(path, _index, snapshot_bytes=None):
            loaded_paths.append(Path(path))
            loaded_bytes.append(snapshot_bytes)
            return Image.new("RGB", (4, 4), "white")

        def mutate_during_ocr(_image, on_delta=None, max_retries=0):
            source.write_bytes(changed)
            return "<p>safe text</p>"

        with (
            mock.patch.object(service.grm_ocr, "probe", return_value=True),
            mock.patch.object(service, "load_page", side_effect=load_snapshot),
            mock.patch.object(
                service.grm_ocr, "ocr_page_raw", side_effect=mutate_during_ocr
            ),
            mock.patch.object(
                service.grm_ocr, "raw_to_markdown", return_value="safe text"
            ),
            mock.patch.object(
                service.grm_ocr, "raw_to_html", return_value="<p>safe text</p>"
            ),
            mock.patch.object(service.time, "sleep", return_value=None),
        ):
            service.process_file(source, inbox)

        self.assertEqual(source.read_bytes(), changed)
        self.assertEqual(len(loaded_paths), 1)
        self.assertNotEqual(loaded_paths[0], source)
        self.assertEqual(loaded_paths[0].name, ".source.png")
        self.assertEqual(loaded_bytes, [original])
        self.assertEqual(list(service.PROCESSED_DIR.glob("*")), [])
        self.assertEqual(service.STATE["history"], [])
        self.assertEqual(list(service.JOBS_DIR.glob("*/.source.*")), [])

    def test_aggregate_ocr_budget_stops_new_pages_and_closes_loaded_images(self):
        inbox = self.root / "data" / "inbox"
        inbox.mkdir(exist_ok=True)
        source = inbox / "bounded.png"
        source.write_bytes(b"\x89PNG\r\n\x1a\n" + b"bounded-payload")
        loaded = []
        remote_calls = []

        class TrackedImage:
            width = 4
            height = 4

            def __init__(self):
                self.closed = False

            def resize(self, _size):
                return self

            def convert(self, _mode):
                return Image.new("RGB", (4, 4), "white")

            def close(self):
                self.closed = True

        def load_bounded(_path, _index, snapshot_bytes=None):
            self.assertEqual(snapshot_bytes, source.read_bytes())
            image = TrackedImage()
            loaded.append(image)
            return image

        def remote(_image, *, on_delta, max_retries):
            remote_calls.append(max_retries)
            on_delta("raw")
            return "raw"

        with (
            mock.patch.object(service, "OCR_CONCURRENCY", 1),
            mock.patch.object(service, "MAX_PAGE_OUTPUT_CHARS", 1000),
            mock.patch.object(service, "MAX_JOB_OUTPUT_CHARS", 1500),
            mock.patch.object(service, "MAX_REMOTE_ATTEMPTS", 4),
            mock.patch.object(service, "count_pages", return_value=4),
            mock.patch.object(service, "load_page", side_effect=load_bounded),
            mock.patch.object(service.grm_ocr, "probe", return_value=True),
            mock.patch.object(service.grm_ocr, "ocr_page_raw", side_effect=remote),
            mock.patch.object(
                service.grm_ocr, "raw_to_markdown", return_value="m" * 400
            ),
            mock.patch.object(
                service.grm_ocr, "raw_to_html", return_value="<p>" + "h" * 400 + "</p>"
            ),
            mock.patch.object(service.time, "sleep", return_value=None),
        ):
            service.process_file(source, inbox)

        self.assertEqual(remote_calls, [0, 0])
        self.assertEqual(len(loaded), 2)
        self.assertTrue(all(image.closed for image in loaded))
        self.assertFalse(source.exists())
        self.assertTrue((service.ON_HOLD_DIR / source.name).is_file())
        self.assertEqual(service.STATE["history"][0]["status"], "cancelled")

    def test_retention_only_deletes_exact_owned_safe_job_directories(self):
        owned = service.JOBS_DIR / "20260810-010101-aaaaaaaa"
        unsafe = service.JOBS_DIR / "20260810-010102-bbbbbbbb"
        arbitrary = service.JOBS_DIR / "customer-archive"
        for directory in (owned, unsafe, arbitrary):
            directory.mkdir(parents=True)
        (owned / "out.txt").write_text("owned", encoding="utf-8")
        unsafe_child = unsafe / "link"
        unsafe_child.write_text("pretend-reparse", encoding="utf-8")
        (arbitrary / "keep.txt").write_text("keep", encoding="utf-8")
        original_check = service._is_reparse_or_symlink

        def mark_unsafe(path):
            return Path(path) == unsafe_child or original_check(Path(path))

        with (
            mock.patch.object(service, "RETENTION_DAYS", 0),
            mock.patch.object(service, "_is_reparse_or_symlink", side_effect=mark_unsafe),
        ):
            removed = service.enforce_retention()
        self.assertEqual(removed, {owned.name})
        self.assertFalse(owned.exists())
        self.assertTrue(unsafe.exists())
        self.assertTrue(arbitrary.exists())

    def test_formula_cells_are_written_as_inert_strings(self):
        output = self.root / "formula.xlsx"
        service.export_xlsx(
            [
                "<p>=WEBSERVICE(&quot;https://example.invalid/text&quot;)</p>"
                "<table><tr><td>=HYPERLINK(&quot;https://example.invalid&quot;,&quot;open&quot;)</td><td>12</td></tr></table>"
            ],
            output,
        )
        workbook = load_workbook(output, data_only=False)
        formula_cell = workbook[workbook.sheetnames[1]]["A1"]
        self.assertEqual(formula_cell.data_type, "s")
        self.assertTrue(formula_cell.value.startswith("'="))
        text_formula = workbook["Text"]["A2"]
        self.assertEqual(text_formula.data_type, "s")
        self.assertTrue(text_formula.value.startswith("'="))

    def test_strict_ocr_allowlist_removes_active_and_remote_markup(self):
        dirty = (
            '<svg><a xlink:href="javascript:alert(1)">x</a></svg>'
            '<form action="/api/cancel"><button>stop</button></form>'
            '<div style="position:fixed" onclick="bad()"><img src="https://tracker.invalid/x">safe</div>'
            '<table><tr><td data-bbox="0 0 10 10">cell</td></tr></table>'
        )
        clean = service.sanitize_ocr_html(dirty)
        for forbidden in ("svg", "form", "button", "style=", "onclick", "img", "tracker.invalid"):
            self.assertNotIn(forbidden, clean.lower())
        self.assertIn("safe", clean)
        self.assertIn("data-bbox", clean)


class EngineConfigurationTest(unittest.TestCase):
    def test_upstream_must_be_explicit_and_plaintext_requires_opt_in(self):
        with mock.patch.dict(os.environ, {}, clear=True):
            with self.assertRaisesRegex(ValueError, "GRM_OCR_API_BASE"):
                service.grm_ocr.load_config()
        with mock.patch.dict(
            os.environ,
            {"GRM_OCR_API_BASE": "http://127.0.0.1:8000/v1"},
            clear=True,
        ):
            with self.assertRaisesRegex(ValueError, "plaintext HTTP"):
                service.grm_ocr.load_config()
        with mock.patch.dict(
            os.environ,
            {
                "GRM_OCR_API_BASE": "http://127.0.0.1:8000/v1",
                "GRM_OCR_ALLOW_INSECURE_HTTP": "1",
                "GRM_OCR_MODEL": "grm-test",
            },
            clear=True,
        ):
            config = service.grm_ocr.load_config()
            self.assertEqual(config.request_timeout, 120)
            self.assertEqual(config.transport_retries, 0)
        with mock.patch.dict(
            os.environ,
            {"GRM_OCR_API_BASE": "https://127.0.0.1:8000/v1"},
            clear=True,
        ):
            with self.assertRaisesRegex(ValueError, "GRM_OCR_MODEL"):
                service.grm_ocr.load_config()

    def test_engine_bounds_are_revalidated(self):
        with mock.patch.dict(
            os.environ,
            {
                "GRM_OCR_API_BASE": "https://127.0.0.1:8000/v1",
                "GRM_OCR_MODEL": "grm-test",
                "GRM_OCR_TIMEOUT": "999",
            },
            clear=True,
        ):
            with self.assertRaisesRegex(ValueError, "GRM_OCR_TIMEOUT"):
                service.grm_ocr.load_config()
        base = {
            "api_key": "profile-secret-value",
            "model": "grm-test",
            "max_tokens": 8192,
            "request_timeout": 120,
            "transport_retries": 0,
            "ca_bundle": None,
            "allow_insecure_http": False,
        }
        with self.assertRaises(ValueError):
            service.grm_ocr.validate_config(
                service.grm_ocr.EngineConfig(
                    api_base="https://engine.invalid/v1\nInjected: value", **base
                )
            )
        with self.assertRaises(ValueError):
            service.grm_ocr.validate_config(
                service.grm_ocr.EngineConfig(
                    api_base="https://engine.invalid/v1",
                    **{**base, "allow_insecure_http": True},
                )
            )
        with self.assertRaises(ValueError):
            service.grm_ocr.validate_config(
                service.grm_ocr.EngineConfig(
                    api_base="https://engine.invalid/v1",
                    **{**base, "model": "model with spaces"},
                )
            )

    def test_profile_engine_config_ignores_global_env_and_keeps_profiles_isolated(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)

            def write_profile(name, api_base, token):
                config_dir = root / name / "document-reader" / "config"
                config_dir.mkdir(parents=True)
                config = {
                    "schema": 1,
                    "api_base": api_base,
                    "model": "grm-test",
                    "max_tokens": 2048,
                    "request_timeout": 30,
                    "transport_retries": 0,
                    "ca_bundle": None,
                    "allow_insecure_http": False,
                    "allow_remote_mcp_ocr": False,
                }
                config_path = config_dir / "engine.json"
                token_path = config_dir / "engine.token"
                atomic_write_json(config_path, config)
                write_private_single_line(
                    token_path, token, minimum=16, maximum=2048
                )
                return config_path, token_path

            first_paths = write_profile(
                "alpha", "https://alpha.invalid/v1", "alpha-secret-value"
            )
            second_paths = write_profile(
                "beta", "https://beta.invalid/v1", "beta-secret-value"
            )
            with mock.patch.dict(
                os.environ,
                {
                    "GRM_OCR_API_BASE": "https://poison.invalid/v1",
                    "GRM_OCR_API_KEY": "global-secret",
                },
                clear=True,
            ):
                first = service.grm_ocr.load_profile_config(*first_paths)
                second = service.grm_ocr.load_profile_config(*second_paths)
                self.assertEqual(first.api_base, "https://alpha.invalid/v1")
                self.assertEqual(first.api_key, "alpha-secret-value")
                self.assertEqual(second.api_base, "https://beta.invalid/v1")
                self.assertEqual(second.api_key, "beta-secret-value")
                self.assertNotIn("alpha-secret-value", repr(first))
                service.grm_ocr.configure(first)
                try:
                    selected = service.grm_ocr.current_config()
                    self.assertEqual(selected.api_base, first.api_base)
                    self.assertEqual(selected.api_key, first.api_key)
                finally:
                    service.grm_ocr.configure(None)

    def test_lifecycle_and_service_engine_bounds_have_exact_parity(self):
        with tempfile.TemporaryDirectory() as directory:
            runtime = resolve_profile_runtime(home=Path(directory), profile_name="default")
            create_profile_directories(runtime)
            lifecycle_engine_config.configure_engine(
                runtime,
                api_base="https://engine.invalid/v1",
                model="grm-test",
                token="profile-secret-value",
                max_tokens=32768,
                request_timeout=300,
                transport_retries=2,
            )
            accepted = service.grm_ocr.load_profile_config(
                runtime.engine_config_file, runtime.engine_token_file
            )
            self.assertEqual(
                (accepted.max_tokens, accepted.request_timeout, accepted.transport_retries),
                (32768, 300, 2),
            )
            normalized, _ = lifecycle_engine_config.validate_engine_config(runtime)
            for field, rejected in (
                ("max_tokens", 32769),
                ("request_timeout", 301),
                ("transport_retries", 3),
            ):
                raw = dict(normalized)
                raw[field] = rejected
                with self.assertRaises(ProfileRuntimeError):
                    lifecycle_engine_config.validate_engine_config(runtime, raw)
                kwargs = {
                    "api_base": accepted.api_base,
                    "api_key": accepted.api_key,
                    "model": accepted.model,
                    "max_tokens": accepted.max_tokens,
                    "request_timeout": accepted.request_timeout,
                    "transport_retries": accepted.transport_retries,
                    "ca_bundle": accepted.ca_bundle,
                    "allow_insecure_http": accepted.allow_insecure_http,
                    "allow_remote_mcp_ocr": accepted.allow_remote_mcp_ocr,
                }
                kwargs[field] = rejected
                with self.assertRaises(ValueError):
                    service.grm_ocr.validate_config(
                        service.grm_ocr.EngineConfig(**kwargs)
                    )

    def test_profile_engine_token_rejects_whitespace_without_normalizing_it(self):
        with tempfile.TemporaryDirectory() as directory:
            runtime = resolve_profile_runtime(home=Path(directory), profile_name="default")
            create_profile_directories(runtime)
            lifecycle_engine_config.configure_engine(
                runtime,
                api_base="https://engine.invalid/v1",
                model="grm-test",
                token="profile-secret-value",
            )
            runtime.engine_token_file.write_text(
                "profile-secret-value \n", encoding="utf-8"
            )
            with self.assertRaisesRegex(ValueError, "exact bounded line"):
                service.grm_ocr.load_profile_config(
                    runtime.engine_config_file, runtime.engine_token_file
                )

    def test_engine_config_set_rejects_mid_load_endpoint_token_rotation(self):
        with tempfile.TemporaryDirectory() as directory:
            runtime = resolve_profile_runtime(home=Path(directory), profile_name="default")
            create_profile_directories(runtime)
            lifecycle_engine_config.configure_engine(
                runtime,
                api_base="https://old.invalid/v1",
                model="grm-test",
                token="old-profile-secret",
            )
            original_read = service.grm_ocr.read_private_value
            rotated = False

            def rotate_after_first_token(path, minimum, maximum):
                nonlocal rotated
                value = original_read(path, minimum, maximum)
                if not rotated:
                    rotated = True
                    lifecycle_engine_config.configure_engine(
                        runtime,
                        api_base="https://new.invalid/v1",
                        model="grm-test",
                        token="new-profile-secret",
                    )
                return value

            with mock.patch.object(
                service.grm_ocr,
                "read_private_value",
                side_effect=rotate_after_first_token,
            ):
                with self.assertRaisesRegex(ValueError, "changed while loading"):
                    service.grm_ocr.load_profile_config(
                        runtime.engine_config_file, runtime.engine_token_file
                    )

    def test_engine_config_set_rejects_mid_load_ca_rotation(self):
        with tempfile.TemporaryDirectory() as directory:
            runtime = resolve_profile_runtime(home=Path(directory), profile_name="default")
            create_profile_directories(runtime)
            ca_path = runtime.config_dir / "engine-ca.pem"
            write_private_single_line(
                ca_path, "old-ca-material", minimum=1, maximum=1024
            )
            lifecycle_engine_config.configure_engine(
                runtime,
                api_base="https://engine.invalid/v1",
                model="grm-test",
                token="profile-secret-value",
                ca_bundle=ca_path.name,
            )
            original_read = service.grm_ocr.read_regular_bytes
            rotated = False

            def rotate_after_first_ca(path, **kwargs):
                nonlocal rotated
                value = original_read(path, **kwargs)
                if Path(path) == ca_path and not rotated:
                    rotated = True
                    write_private_single_line(
                        ca_path, "new-ca-material", minimum=1, maximum=1024
                    )
                return value

            with mock.patch.object(
                service.grm_ocr,
                "read_regular_bytes",
                side_effect=rotate_after_first_ca,
            ):
                with self.assertRaisesRegex(ValueError, "changed while loading"):
                    service.grm_ocr.load_profile_config(
                        runtime.engine_config_file, runtime.engine_token_file
                    )

    @unittest.skipUnless(os.name == "nt", "Windows ACL validation")
    def test_profile_engine_token_rejects_a_foreign_windows_acl_principal(self):
        with tempfile.TemporaryDirectory() as directory:
            runtime = resolve_profile_runtime(home=Path(directory), profile_name="default")
            create_profile_directories(runtime)
            lifecycle_engine_config.configure_engine(
                runtime,
                api_base="https://engine.invalid/v1",
                model="grm-test",
                token="profile-secret-value",
            )
            result = subprocess.run(
                [
                    "icacls",
                    str(runtime.engine_token_file),
                    "/grant",
                    "*S-1-5-32-545:(R)",
                ],
                capture_output=True,
                text=True,
                timeout=15,
                check=False,
            )
            self.assertEqual(result.returncode, 0)
            try:
                with self.assertRaisesRegex(PermissionError, "foreign principal"):
                    service.grm_ocr.load_profile_config(
                        runtime.engine_config_file, runtime.engine_token_file
                    )
                with self.assertRaisesRegex(
                    ProfileRuntimeError, "foreign or inherited principal"
                ):
                    _harden_windows_secret_acl(runtime.engine_token_file)
            finally:
                # The fixture intentionally has a foreign explicit ACE.  Production
                # must refuse to overwrite it; this test owns the exact temporary
                # file and removes it instead of asking the hardener to normalize it.
                runtime.engine_token_file.unlink(missing_ok=True)


class ServiceConfigurationTest(unittest.TestCase):
    def test_config_is_profile_owned_and_exact(self):
        with tempfile.TemporaryDirectory() as directory:
            home = Path(directory).resolve()
            plugin = home / "document-reader"
            data = plugin / "data"
            release = plugin / "runtime" / "releases" / "0.1.0-test"
            config_dir = plugin / "config"
            for path in (
                data / "inbox", data / "processed", data / "jobs", data / "state",
                data / "logs", release / "install", release / ".venv" / "bin", config_dir,
            ):
                path.mkdir(parents=True, exist_ok=True)
            entry = release / "install" / "profile_service.py"
            python = release / ".venv" / "bin" / "python"
            entry.write_text("# service", encoding="utf-8")
            python.write_text("python", encoding="utf-8")
            fingerprint = hashlib.sha256(
                (os.path.normcase(str(home)) if os.name == "nt" else str(home)).encode()
            ).hexdigest()
            owner = hashlib.sha256(f"document-reader\0{fingerprint}".encode()).hexdigest()
            config = {
                "schema": 1,
                "plugin": "document-reader",
                "version": "0.1.0",
                "api_version": 1,
                "profile": "default",
                "profile_fingerprint": fingerprint,
                "owner_id": owner,
                "instance_id": "c" * 32,
                "hermes_home": str(home),
                "plugin_root": str(plugin),
                "data_root": str(data),
                "inbox": str(data / "inbox"),
                "processed": str(data / "processed"),
                "jobs": str(data / "jobs"),
                "state": str(data / "state"),
                "logs": str(data / "logs"),
                "bind": "127.0.0.1",
                "port": 28000,
                "token_file": str(config_dir / "service.token"),
                "release_id": "0.1.0-test",
                "release_root": str(release),
                "service_entry": str(entry),
                "runtime_python": str(python),
                "task_name": f"Hermes_DocumentReader_{fingerprint[:12]}",
            }
            path = config_dir / "service.json"
            path.write_text(json.dumps(config), encoding="utf-8")
            loaded = service.load_service_config(path)
            self.assertEqual(loaded["owner_id"], owner)
            original_open = service.grm_ocr._open_no_follow
            replaced = False

            def replace_before_open(candidate):
                nonlocal replaced
                if Path(candidate) == path and not replaced:
                    replaced = True
                    payload = path.read_bytes()
                    path.unlink()
                    path.write_bytes(payload)
                return original_open(candidate)

            with mock.patch.object(
                service.grm_ocr, "_open_no_follow", side_effect=replace_before_open
            ):
                with self.assertRaisesRegex(ValueError, "changed identity"):
                    service.load_service_config(path)
            config["owner_id"] = "0" * 64
            path.write_text(json.dumps(config), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "owner"):
                service.load_service_config(path)


if __name__ == "__main__":
    unittest.main()
