# -*- coding: utf-8 -*-
"""Hermes Document Reader watched-folder OCR service with a live web UI.

Profile usage:
  1. Drop a scanned PDF into the selected profile's private inbox, or drag it
     onto the authenticated web page.
  2. Watch it scan through the profile-scoped loopback service.
  3. Outputs land in the selected profile's retained processed directory.

Engine: bounded document rendering plus the profile's configured
OpenAI-compatible multimodal OCR endpoint (thinking disabled, repeat retry,
and output normalization).

Run through the profile lifecycle wrapper with an owned service.json.

All live job state is kept in memory and served as JSON (/api/state) — no
status-file writes, which kills the whole Windows file-lock class the
single-shot viewer had to retry around.
"""

import argparse
import base64
import copy
import dataclasses
import hashlib
import hmac
import io
import ipaddress
import json
import math
import os
import re
import secrets
import shutil
import socket
import stat
import sys
import tempfile
import threading
import time
import urllib.parse
from concurrent.futures import ThreadPoolExecutor
from http.cookies import CookieError, SimpleCookie
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
ENGINE_DIR = PROJECT_ROOT / "engine"
sys.path.insert(0, str(ENGINE_DIR if ENGINE_DIR.is_dir() else PROJECT_ROOT))

import pypdfium2 as pdfium
from chandra.input import flatten, load_file

import grm_ocr

VIEWER_DIR = Path(__file__).parent
VERSION = "0.1.0"
SERVICE_NAME = "hermes-document-reader"
PLUGIN_ID = "document-reader"
API_VERSION = 1
PROFILE_ID = "default"
DATA_ROOT = Path(tempfile.gettempdir()) / "hermes-document-reader-unconfigured"
STATE_DIR = DATA_ROOT
JOBS_DIR = DATA_ROOT / "jobs"
PROCESSED_DIR = DATA_ROOT / "processed"
ON_HOLD_DIR = DATA_ROOT / "on-hold"
NEEDS_REVIEW_DIR = DATA_ROOT / "needs-review"
QUARANTINE_DIR = DATA_ROOT / "quarantine"
HISTORY_PATH = STATE_DIR / "history.json"
_RUNTIME_CHILD_IDENTITIES = {}
DISPLAY_WIDTH = 1100
SUPPORTED = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".bmp"}
# Concurrent page requests to the configured OCR endpoint. Three balances
# document throughput with bounded decoded-image memory and endpoint load.
try:
    OCR_CONCURRENCY = int(os.environ.get("OCR_CONCURRENCY", "3"))
except ValueError as exc:
    raise ValueError("OCR_CONCURRENCY must be an integer between 1 and 4") from exc
if not 1 <= OCR_CONCURRENCY <= 4:
    raise ValueError("OCR_CONCURRENCY must be between 1 and 4")
MAX_UPLOAD_BYTES = 100 * 1024 * 1024
MAX_PAGES = 100
MAX_IMAGE_PIXELS = 25_000_000
MAX_IMAGE_DIMENSION = 12_000
MAX_CONCURRENT_DECODED_BYTES = 256 * 1024 * 1024
MAX_PAGE_OUTPUT_CHARS = 250_000
MAX_JOB_OUTPUT_CHARS = 4_000_000
MAX_REMOTE_ATTEMPTS = 100
MAX_PAGE_OCR_SECONDS = 120
MAX_FILE_RETRIES = 3
RETRY_DELAY_SECONDS = 300
RETENTION_DAYS = 30
MAX_RETAINED_JOB_BYTES = 2 * 1024 * 1024 * 1024
MAX_RETAINED_JOBS = 200
MAX_HISTORY_BYTES = 2 * 1024 * 1024
MAX_REQUEST_PATH = 2048
MAX_HEADER_COUNT = 64
MAX_HEADER_BYTES = 32768
MAX_HEADER_VALUE = 8192
REQUEST_SOCKET_TIMEOUT = 20
MAX_HTTP_THREADS = 16
MAX_INLINE_HTML_BYTES = 4 * 1024 * 1024
MAX_WATCHER_TRACKED = 2048
JOB_ID_PATTERN = re.compile(r"\A\d{8}-\d{6}-[0-9a-f]{8}\Z")
JOB_DOWNLOAD_SUFFIXES = {".jpg", ".md", ".txt", ".html", ".xlsx"}

LOCK = threading.Lock()
STATE = {
    "queue": [],       # [{name, size}] waiting
    "job": None,       # live job dict (same shape the viewer UI expects) + partial
    "history": [],     # [{name, when, pages, secs, chars, links:{md,xlsx}, errors}]
}
_pending_paths = []    # Path objects matching STATE["queue"]
_retry_after = {}      # path str -> epoch; failed files wait this long before requeue
_retry_counts = {}     # path str -> bounded attempts before quarantine
_RUNTIME_IDENTITY = None

SERVICE_CONFIG_KEYS = {
    "schema", "plugin", "version", "api_version", "profile",
    "profile_fingerprint", "owner_id", "instance_id", "hermes_home",
    "plugin_root", "data_root", "inbox", "processed", "jobs", "state",
    "logs", "bind", "port", "token_file", "release_id", "release_root",
    "service_entry", "runtime_python", "task_name",
}


def bounded_int(value, name: str, minimum: int, maximum: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{name} must be an integer") from exc
    if not minimum <= parsed <= maximum:
        raise ValueError(f"{name} must be between {minimum} and {maximum}")
    return parsed


def normalize_profile(profile: str) -> str:
    profile = str(profile or "default").strip()
    if not re.fullmatch(r"[a-z0-9][a-z0-9_-]{0,63}", profile):
        raise ValueError(
            "profile must be canonical lowercase Hermes syntax: letters, "
            "numbers, underscore, or dash"
        )
    return profile


def default_hermes_home() -> Path:
    configured = os.environ.get("HERMES_HOME", "").strip()
    if configured:
        return Path(configured).expanduser()
    try:
        from hermes_constants import get_hermes_home

        discovered = get_hermes_home()
        if discovered:
            return Path(discovered).expanduser()
    except (ImportError, OSError, TypeError, ValueError):
        pass
    if os.name == "nt":
        base = os.environ.get("LOCALAPPDATA", "").strip()
        if base:
            return Path(base) / "hermes"
    return Path.home() / ".hermes"


def default_data_root(profile: str) -> Path:
    configured = os.environ.get("DOCUMENT_READER_DATA_ROOT", "").strip()
    if configured:
        return Path(configured).expanduser()
    normalize_profile(profile)
    return default_hermes_home() / "document-reader" / "data"


def _private_directory(path: Path) -> None:
    path = Path(os.path.abspath(os.fspath(path)))
    grm_ocr._reject_reparse_chain(path.parent, "private directory")
    path.mkdir(parents=True, exist_ok=True)
    grm_ocr._reject_reparse_chain(path, "private directory")
    info = path.lstat()
    if not stat.S_ISDIR(info.st_mode) or grm_ocr._is_link_or_reparse_info(info):
        raise ValueError("private directory must be a regular non-reparse directory")
    try:
        path.chmod(0o700)
    except OSError:
        pass


def configure_runtime(
    profile: str,
    data_root: Path,
    *,
    jobs_root: Path | None = None,
    state_root: Path | None = None,
    processed_root: Path | None = None,
) -> Path:
    """Select one profile's private runtime root, always outside source."""
    global PROFILE_ID, DATA_ROOT, STATE_DIR, JOBS_DIR, PROCESSED_DIR
    global ON_HOLD_DIR, NEEDS_REVIEW_DIR, QUARANTINE_DIR, HISTORY_PATH
    global _RUNTIME_CHILD_IDENTITIES, _RUNTIME_IDENTITY
    profile = normalize_profile(profile)
    root = Path(os.path.abspath(os.fspath(Path(data_root).expanduser())))
    project = PROJECT_ROOT.resolve()
    if root == project or root.is_relative_to(project):
        raise ValueError("document-reader data root must be outside the source/install tree")
    _private_directory(root)
    jobs = Path(os.path.abspath(os.fspath(Path(jobs_root or root / "jobs").expanduser())))
    state_root = Path(os.path.abspath(os.fspath(Path(state_root or root / "state").expanduser())))
    processed = Path(os.path.abspath(os.fspath(Path(processed_root or root / "processed").expanduser())))
    for label, path in (("jobs", jobs), ("state", state_root), ("processed", processed)):
        if not path.is_relative_to(root):
            raise ValueError(f"{label} root must be inside the profile data root")
        _private_directory(path)
    PROFILE_ID = profile
    DATA_ROOT = root
    STATE_DIR = state_root
    JOBS_DIR = jobs
    PROCESSED_DIR = processed
    ON_HOLD_DIR = root / "on-hold"
    NEEDS_REVIEW_DIR = root / "needs-review"
    QUARANTINE_DIR = root / "quarantine"
    for path in (ON_HOLD_DIR, NEEDS_REVIEW_DIR, QUARANTINE_DIR):
        _private_directory(path)
    HISTORY_PATH = STATE_DIR / "history.json"
    _RUNTIME_CHILD_IDENTITIES = {}
    _RUNTIME_IDENTITY = None
    return root


def data_root_fingerprint(root: Path | None = None) -> str:
    canonical = os.path.normcase(str((root or DATA_ROOT).resolve()))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:16]


def strong_token(token: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z0-9_-]{43,128}", str(token or "")))


def _canonical(path) -> Path:
    return Path(path).expanduser().resolve(strict=False)


def _contained(base: Path, child: Path, label: str) -> Path:
    base = _canonical(base)
    child = _canonical(child)
    if not child.is_relative_to(base):
        raise ValueError(f"service config {label} escapes {base}")
    return child


def load_service_config(path: Path) -> dict:
    path = Path(path)
    if not path.is_absolute() or path.name != "service.json":
        raise ValueError("--config must be an absolute regular service.json path")
    try:
        value = json.loads(
            grm_ocr.read_regular_bytes(
                path, maximum=64 * 1024, label="service config"
            ).decode("utf-8")
        )
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("service config must contain valid UTF-8 JSON") from exc
    if not isinstance(value, dict) or set(value) != SERVICE_CONFIG_KEYS:
        raise ValueError("service config has an unexpected schema")
    expected_scalars = {
        "schema": 1,
        "plugin": PLUGIN_ID,
        "version": VERSION,
        "api_version": API_VERSION,
        "bind": "127.0.0.1",
    }
    for key, expected in expected_scalars.items():
        if value.get(key) != expected:
            raise ValueError(f"service config {key} does not match this release")
    value["profile"] = normalize_profile(value.get("profile"))
    for key in ("profile_fingerprint", "owner_id"):
        if not re.fullmatch(r"[0-9a-f]{64}", str(value.get(key, ""))):
            raise ValueError(f"service config {key} must be 64 lowercase hex characters")
    if not re.fullmatch(r"[0-9a-f]{32}", str(value.get("instance_id", ""))):
        raise ValueError("service config instance_id must be 32 lowercase hex characters")
    if not re.fullmatch(r"[0-9A-Za-z._-]{1,96}", str(value.get("release_id", ""))):
        raise ValueError("service config release_id is invalid")
    value["port"] = bounded_int(value.get("port"), "service config port", 1024, 65535)

    home = _canonical(value["hermes_home"])
    fingerprint = hashlib.sha256(
        (os.path.normcase(str(home)) if os.name == "nt" else str(home)).encode("utf-8")
    ).hexdigest()
    if not hmac.compare_digest(fingerprint, value["profile_fingerprint"]):
        raise ValueError("service config profile fingerprint does not match HERMES_HOME")
    owner = hashlib.sha256(f"{PLUGIN_ID}\0{fingerprint}".encode("utf-8")).hexdigest()
    if not hmac.compare_digest(owner, value["owner_id"]):
        raise ValueError("service config owner does not match the selected profile")

    plugin_root = _canonical(value["plugin_root"])
    if plugin_root != home / PLUGIN_ID:
        raise ValueError("service config plugin_root is not owned by the selected profile")
    if Path(os.path.abspath(path)) != plugin_root / "config" / "service.json":
        raise ValueError("service config path is not owned by the selected profile")
    data_root = _canonical(value["data_root"])
    if data_root != plugin_root / "data":
        raise ValueError("service config data_root is not the profile plugin data root")
    exact_paths = {
        "inbox": data_root / "inbox",
        "processed": data_root / "processed",
        "jobs": data_root / "jobs",
        "state": data_root / "state",
        "logs": data_root / "logs",
        "token_file": plugin_root / "config" / "service.token",
    }
    for key, expected in exact_paths.items():
        if _canonical(value[key]) != expected:
            raise ValueError(f"service config {key} is not the owned profile path")
    release_root = _contained(plugin_root / "runtime" / "releases", value["release_root"], "release_root")
    if release_root.name != value["release_id"]:
        raise ValueError("service config release root and release id disagree")
    for key in ("service_entry", "runtime_python"):
        configured = _contained(release_root, value[key], key)
        if not configured.is_file():
            raise ValueError(f"configured {key} does not exist")
    if not re.fullmatch(r"Hermes_DocumentReader_[0-9a-f]{12}", str(value["task_name"])):
        raise ValueError("service config task_name is invalid")
    return value


def read_service_token(path: Path) -> str:
    path = Path(path)
    token = grm_ocr.read_private_value(path, minimum=43, maximum=128)
    if not strong_token(token):
        raise ValueError("service token must be one bounded base64url value")
    return token


def capture_runtime_identity(data_root: Path, owner_id: str, profile: str):
    root = Path(os.path.abspath(os.fspath(data_root)))
    grm_ocr._reject_reparse_chain(root, "profile data root")
    info = root.lstat()
    if not stat.S_ISDIR(info.st_mode) or grm_ocr._is_link_or_reparse_info(info):
        raise RuntimeError("profile data root is no longer a regular owned directory")
    return (
        os.path.normcase(str(root)) if os.name == "nt" else str(root),
        int(info.st_dev),
        int(info.st_ino),
        int(getattr(info, "st_uid", 0)),
        str(owner_id),
        str(profile),
    )


def capture_runtime_directories(directories: dict[str, Path]) -> dict[str, tuple]:
    captured = {}
    root = Path(os.path.abspath(os.fspath(DATA_ROOT)))
    for label, raw_path in directories.items():
        path = Path(os.path.abspath(os.fspath(raw_path)))
        if path == root or not path.is_relative_to(root):
            raise RuntimeError(f"runtime {label} directory escapes the profile data root")
        grm_ocr._reject_reparse_chain(path, f"runtime {label} directory")
        info = path.lstat()
        if not stat.S_ISDIR(info.st_mode) or grm_ocr._is_link_or_reparse_info(info):
            raise RuntimeError(f"runtime {label} directory is not a regular directory")
        key = os.path.normcase(str(path)) if os.name == "nt" else str(path)
        captured[key] = (
            label,
            int(info.st_dev),
            int(info.st_ino),
            int(getattr(info, "st_uid", 0)),
        )
    return captured


def revalidate_runtime_directory(path: Path) -> Path:
    revalidate_runtime_identity(_RUNTIME_IDENTITY)
    literal = Path(os.path.abspath(os.fspath(path)))
    if _RUNTIME_IDENTITY is None and not _RUNTIME_CHILD_IDENTITIES:
        grm_ocr._reject_reparse_chain(literal, "runtime directory")
        info = literal.lstat()
        if not stat.S_ISDIR(info.st_mode) or grm_ocr._is_link_or_reparse_info(info):
            raise RuntimeError("runtime directory is not a regular directory")
        return literal
    root = Path(os.path.abspath(os.fspath(DATA_ROOT)))
    if literal == root or not literal.is_relative_to(root):
        raise RuntimeError("runtime directory escapes the profile data root")
    grm_ocr._reject_reparse_chain(literal, "runtime directory")
    info = literal.lstat()
    if not stat.S_ISDIR(info.st_mode) or grm_ocr._is_link_or_reparse_info(info):
        raise RuntimeError("runtime directory is not a regular directory")
    if _RUNTIME_CHILD_IDENTITIES:
        key = os.path.normcase(str(literal)) if os.name == "nt" else str(literal)
        expected = _RUNTIME_CHILD_IDENTITIES.get(key)
        if expected is None or (
            int(info.st_dev), int(info.st_ino), int(getattr(info, "st_uid", 0))
        ) != expected[1:]:
            raise RuntimeError("runtime child-directory ownership changed")
    return literal


def revalidate_runtime_identity(expected=None) -> None:
    expected = expected or _RUNTIME_IDENTITY
    if expected is None:
        return
    current = capture_runtime_identity(DATA_ROOT, expected[4], expected[5])
    if current != expected or PROFILE_ID != expected[5]:
        raise RuntimeError("profile data-root ownership changed while the service was running")
    for path in tuple(_RUNTIME_CHILD_IDENTITIES):
        literal = Path(path)
        grm_ocr._reject_reparse_chain(literal, "runtime directory")
        info = literal.lstat()
        expected_child = _RUNTIME_CHILD_IDENTITIES[path]
        if (
            not stat.S_ISDIR(info.st_mode)
            or grm_ocr._is_link_or_reparse_info(info)
            or (
                int(info.st_dev), int(info.st_ino), int(getattr(info, "st_uid", 0))
            ) != expected_child[1:]
        ):
            raise RuntimeError("profile child-directory ownership changed")


class RuntimeOwnerLock:
    def __init__(self, path: Path):
        self.path = Path(path)
        self.handle = None

    def acquire(self):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.handle = self.path.open("a+b", buffering=0)
        if self.path.stat().st_size == 0:
            self.handle.write(b"0")
            self.handle.flush()
        self.handle.seek(0)
        try:
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(self.handle.fileno(), msvcrt.LK_NBLCK, 1)
            else:
                import fcntl

                fcntl.flock(self.handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        except (OSError, BlockingIOError) as exc:
            self.handle.close()
            self.handle = None
            raise RuntimeError("another Document Reader owns this profile runtime") from exc
        return self

    def release(self):
        if self.handle is None:
            return
        try:
            self.handle.seek(0)
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(self.handle.fileno(), msvcrt.LK_UNLCK, 1)
            else:
                import fcntl

                fcntl.flock(self.handle.fileno(), fcntl.LOCK_UN)
        finally:
            self.handle.close()
            self.handle = None

    def __enter__(self):
        return self.acquire()

    def __exit__(self, *_):
        self.release()


def is_loopback_bind(bind: str) -> bool:
    if str(bind).lower() == "localhost":
        return True
    try:
        return ipaddress.ip_address(bind).is_loopback
    except ValueError:
        return False


def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)


def classified_error(error: Exception | str) -> str:
    """Map internal failures to a small path/URL/secret-free public vocabulary."""
    if isinstance(error, (TimeoutError, socket.timeout)):
        return "operation timed out"
    kind = type(error).__name__.lower()
    module = type(error).__module__.lower()
    lowered = str(error).lower()
    if "timeout" in kind or "timeout" in lowered:
        return "operation timed out"
    if any(word in lowered for word in ("source changed", "snapshot", "changed during ocr")):
        return "document changed during processing"
    if module.startswith(("httpx", "httpcore", "openai")) or any(
        word in kind for word in ("connection", "transport", "http")
    ):
        return "OCR service request failed"
    if isinstance(error, ValueError):
        return "document failed safety validation"
    if isinstance(error, (PermissionError, OSError)):
        return "private storage operation failed"
    return "document processing failed"


def _strict_request_target(value: str) -> urllib.parse.SplitResult:
    """Parse one origin-form request target and decode its route exactly once."""

    try:
        parsed = urllib.parse.urlsplit(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("invalid request target") from exc
    if parsed.scheme or parsed.netloc or not parsed.path.startswith("/"):
        raise ValueError("invalid request target")
    for component in (parsed.path, parsed.query, parsed.fragment):
        cursor = 0
        while True:
            cursor = component.find("%", cursor)
            if cursor < 0:
                break
            escape = component[cursor + 1 : cursor + 3]
            if len(escape) != 2 or any(ch not in "0123456789abcdefABCDEF" for ch in escape):
                raise ValueError("invalid request target escape")
            cursor += 3
    try:
        route = urllib.parse.unquote_to_bytes(parsed.path).decode("utf-8", errors="strict")
    except UnicodeDecodeError as exc:
        raise ValueError("invalid request target encoding") from exc
    if (
        not route.startswith("/")
        or "\\" in route
        or any(ord(ch) < 0x20 or ord(ch) == 0x7F for ch in route)
    ):
        raise ValueError("invalid request target route")
    return parsed._replace(path=route)


def _history_download(job_id: str, value, suffix: str) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str) or len(value) > 400:
        raise ValueError("history link is invalid")
    parsed = urllib.parse.urlsplit(value)
    decoded = urllib.parse.unquote(parsed.path)
    prefix = f"/jobs/{job_id}/"
    filename = decoded[len(prefix) :] if decoded.startswith(prefix) else ""
    if (
        parsed.scheme
        or parsed.netloc
        or parsed.query
        or parsed.fragment
        or not filename
        or "/" in filename
        or "\\" in filename
        or Path(filename).name != filename
        or not filename.lower().endswith(suffix)
    ):
        raise ValueError("history link is invalid")
    return value


def normalize_history_entry(entry) -> dict:
    if not isinstance(entry, dict):
        raise ValueError("history entry must be an object")
    job_id = entry.get("id")
    if not isinstance(job_id, str) or not JOB_ID_PATTERN.fullmatch(job_id):
        raise ValueError("history job id is invalid")
    raw_name = entry.get("name")
    if not isinstance(raw_name, str) or not raw_name or len(raw_name) > 240:
        raise ValueError("history document name is invalid")
    name = sanitize_name(raw_name)[:180]
    when = entry.get("when")
    if not isinstance(when, str) or not re.fullmatch(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}", when):
        raise ValueError("history timestamp is invalid")
    status = entry.get("status")
    if status not in {"finished", "finished_with_errors", "cancelled"}:
        raise ValueError("history status is invalid")

    def integer(key, maximum):
        value = entry.get(key)
        if type(value) is not int or not 0 <= value <= maximum:
            raise ValueError(f"history {key} is invalid")
        return value

    seconds = entry.get("secs")
    if isinstance(seconds, bool) or not isinstance(seconds, (int, float)):
        raise ValueError("history secs is invalid")
    seconds = float(seconds)
    if not 0 <= seconds <= 366 * 86400:
        raise ValueError("history secs is invalid")
    links = entry.get("links")
    if not isinstance(links, dict):
        raise ValueError("history links are invalid")
    return {
        "id": job_id,
        "name": name,
        "status": status,
        "when": when,
        "pages": integer("pages", MAX_PAGES),
        "errors": integer("errors", MAX_PAGES),
        "secs": seconds,
        "chars": integer("chars", 1_000_000_000),
        "links": {
            "md": _history_download(job_id, links.get("md"), ".txt"),
            "xlsx": _history_download(job_id, links.get("xlsx"), ".xlsx"),
        },
    }


def load_history():
    revalidate_runtime_directory(STATE_DIR)
    if HISTORY_PATH.exists():
        try:
            loaded = json.loads(
                grm_ocr.read_regular_bytes(
                    HISTORY_PATH,
                    maximum=MAX_HISTORY_BYTES,
                    label="history",
                ).decode("utf-8")
            )
            if not isinstance(loaded, list):
                raise ValueError("history root must be a list")
            STATE["history"] = [
                normalize_history_entry(entry)
                for entry in loaded[:MAX_RETAINED_JOBS]
            ]
        except (OSError, UnicodeDecodeError, json.JSONDecodeError, ValueError) as exc:
            quarantine = HISTORY_PATH.with_name(
                f"history.corrupt-{time.strftime('%Y%m%d-%H%M%S')}-{secrets.token_hex(3)}.json"
            )
            try:
                os.replace(HISTORY_PATH, quarantine)
            except OSError as move_error:
                raise RuntimeError(
                    "history is invalid and could not be quarantined; refusing to overwrite it"
                ) from move_error
            STATE["history"] = []
            log(f"invalid history quarantined as {quarantine.name}")


def atomic_write_bytes(path: Path, data: bytes) -> None:
    grm_ocr._reject_reparse_chain(path.parent, "atomic output directory")
    if not path.parent.is_dir() or _is_reparse_or_symlink(path.parent):
        raise RuntimeError("atomic output directory is unsafe")
    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb", dir=path.parent, prefix=f".{path.name}-", suffix=".tmp", delete=False
        ) as handle:
            temp_path = Path(handle.name)
            handle.write(data)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_path, path)
        temp_path = None
        grm_ocr._reject_reparse_chain(path.parent, "atomic output directory")
        try:
            directory_fd = os.open(path.parent, os.O_RDONLY)
        except OSError:
            directory_fd = None
        if directory_fd is not None:
            try:
                os.fsync(directory_fd)
            except OSError:
                pass
            finally:
                os.close(directory_fd)
    finally:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)


def save_history():
    revalidate_runtime_directory(STATE_DIR)
    normalized = [normalize_history_entry(entry) for entry in STATE["history"]]
    STATE["history"] = normalized
    payload = json.dumps(normalized, indent=1).encode("utf-8")
    if len(payload) > MAX_HISTORY_BYTES:
        raise RuntimeError("history payload exceeds the bounded storage limit")
    atomic_write_bytes(HISTORY_PATH, payload)


def sanitize_name(name: str) -> str:
    name = Path(name).name
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip() or "upload.pdf"


def sanitize_ocr_html(html: str) -> str:
    return grm_ocr.sanitize_ocr_html(html)


def extract_regions(raw: str, limit: int = 48) -> list:
    """Extract normalized layout boxes from GRM's streamed HTML response."""
    soup = BeautifulSoup(raw or "", "html.parser")
    regions = []
    for tag in soup.find_all(attrs={"data-bbox": True}):
        try:
            x0, y0, x1, y1 = (float(v) for v in tag["data-bbox"].split())
        except (TypeError, ValueError):
            continue
        x0, y0 = max(0.0, min(1000.0, x0)), max(0.0, min(1000.0, y0))
        x1, y1 = max(0.0, min(1000.0, x1)), max(0.0, min(1000.0, y1))
        if x1 <= x0 or y1 <= y0:
            continue
        label = str(tag.get("data-label") or (tag.get("class") or [tag.name])[0])
        label_key = label.lower()
        if "table" in label_key or "form" in label_key:
            kind = "data"
        elif "header" in label_key or "section" in label_key:
            kind = "section"
        elif label_key in {"image", "figure", "diagram", "complex-block"}:
            kind = "visual"
        else:
            kind = "text"
        regions.append({
            "x": round(x0 / 10, 2),
            "y": round(y0 / 10, 2),
            "w": round((x1 - x0) / 10, 2),
            "h": round((y1 - y0) / 10, 2),
            "kind": kind,
            "label": label[:32],
        })
    if len(regions) <= limit:
        return regions
    # Keep section/table/form boxes even after a text-heavy page exceeds the
    # live payload cap, then fill the remaining slots with the newest text.
    keep = {i for i, region in enumerate(regions) if region["kind"] != "text"}
    if len(keep) > limit:
        keep = set(sorted(keep)[-limit:])
    for i in range(len(regions) - 1, -1, -1):
        if len(keep) >= limit:
            break
        keep.add(i)
    return [region for i, region in enumerate(regions) if i in keep]


def unique_path(directory: Path, name: str) -> Path:
    p = directory / name
    stem, suffix = p.stem, p.suffix
    k = 1
    while p.exists():
        p = directory / f"{stem} ({k}){suffix}"
        k += 1
    return p


# ---------------------------------------------------------------- excel export

def _table_rows(tbl) -> list:
    rows = []
    for tr in tbl.find_all("tr"):
        rows.append([c.get_text(" ", strip=True) for c in tr.find_all(["td", "th"])])
    return rows


def safe_spreadsheet_text(value: str) -> str:
    """Keep OCR-controlled text inert when a workbook is opened."""
    rendered = str(value)
    if rendered.lstrip(" \t\r\n").startswith(("=", "+", "-", "@")):
        return "'" + rendered
    return rendered


def set_inert_text_cell(worksheet, row: int, column: int, value: str):
    cell = worksheet.cell(row=row, column=column, value=safe_spreadsheet_text(value))
    cell.data_type = "s"
    return cell


def export_xlsx(page_htmls: list, out_path: Path) -> None:
    """Tables from every page → one sheet per unique table; text → a Text sheet.

    Scanned tax forms often carry several identical copies of the same form on
    one page (W-2 Copy B/C/2…) — identical tables are deduplicated so the
    workbook isn't a wall of clones. Trivial 1-cell tables are skipped.
    """
    wb = Workbook()
    text_ws = wb.active
    text_ws.title = "Text"
    text_ws.column_dimensions["A"].width = 110
    row = 1
    seen = set()
    for page_num, html in enumerate(page_htmls, 1):
        soup = BeautifulSoup(html or "", "html.parser")
        page_tables = 0
        dupes = 0
        for tbl in soup.find_all("table"):
            rows = _table_rows(tbl)
            tbl.decompose()
            cells = [c for r in rows for c in r if c]
            if len(cells) < 2:
                continue
            key = repr(rows)
            if key in seen:
                dupes += 1
                continue
            seen.add(key)
            page_tables += 1
            ws = wb.create_sheet(f"P{page_num} Table {page_tables}"[:31])
            widths = {}
            for r, cols in enumerate(rows, 1):
                for c, val in enumerate(cols, 1):
                    num = val.replace(",", "").replace("$", "").strip()
                    if re.fullmatch(r"-?\d+(\.\d+)?", num or "x"):
                        ws.cell(row=r, column=c, value=float(num))
                    else:
                        cell = ws.cell(row=r, column=c, value=safe_spreadsheet_text(val))
                        cell.data_type = "s"
                    widths[c] = min(60, max(widths.get(c, 10), len(val) + 2))
            for c, w in widths.items():
                ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w
        text = soup.get_text("\n", strip=True)
        header = f"--- Page {page_num} ---" + (
            f"  ({dupes} duplicate form copy(ies) omitted)" if dupes else ""
        )
        set_inert_text_cell(text_ws, row, 1, header)
        row += 1
        for line in text.splitlines():
            if line.strip():
                set_inert_text_cell(text_ws, row, 1, line.strip())
                row += 1
        row += 1
    wb.save(out_path)


def export_txt(page_htmls: list) -> str:
    """Readable plain text: document order preserved, tables as aligned rows,
    duplicate form copies noted once — no raw HTML."""
    out = []
    seen = set()
    for page_num, html in enumerate(page_htmls, 1):
        out.append(f"========== Page {page_num} ==========\n")
        soup = BeautifulSoup(html or "", "html.parser")
        for tbl in soup.find_all("table"):
            rows = _table_rows(tbl)
            key = repr(rows)
            if key in seen:
                tbl.replace_with(soup.new_string("[duplicate copy of the table above omitted]\n"))
                continue
            seen.add(key)
            widths = {}
            for cols in rows:
                for i, v in enumerate(cols):
                    widths[i] = min(44, max(widths.get(i, 0), len(v)))
            lines = []
            for cols in rows:
                if not any(c.strip() for c in cols):
                    continue
                lines.append("  ".join(v.ljust(widths[i]) for i, v in enumerate(cols)).rstrip())
            tbl.replace_with(soup.new_string("\n" + "\n".join(lines) + "\n"))
        text = soup.get_text("\n")
        text = re.sub(r"\n{3,}", "\n\n", text).strip()
        out.append(text + "\n")
    return "\n".join(out)


# ---------------------------------------------------------------- job worker

class _Cancelled(Exception):
    pass


# PDFium is NOT thread-safe — concurrent page rendering from worker threads
# dies with a native access violation (exit 255, no traceback). Serialize
# every pdfium touch; page rasterization is ~100ms, so the lock is cheap
# next to the multi-second OCR calls it feeds.
_PDFIUM_LOCK = threading.Lock()


def count_pages(path: Path, snapshot_bytes: bytes | None = None) -> int:
    if path.suffix.lower() == ".pdf":
        with _PDFIUM_LOCK:
            doc = pdfium.PdfDocument(
                snapshot_bytes if snapshot_bytes is not None else str(path)
            )
            try:
                pages = len(doc)
            finally:
                doc.close()
        if not 1 <= pages <= MAX_PAGES:
            raise ValueError(f"document has {pages} pages; limit is {MAX_PAGES}")
        return pages
    return 1


def load_page(path: Path, i: int, snapshot_bytes: bytes | None = None):
    """Load ONE page image. chandra's page_range is 0-indexed. Keeps memory
    at O(workers) instead of O(pages) — a 279-page scanner dump stays flat."""
    from chandra.settings import settings as chandra_settings

    if path.suffix.lower() == ".pdf":
        with _PDFIUM_LOCK:
            document = pdfium.PdfDocument(
                snapshot_bytes if snapshot_bytes is not None else str(path)
            )
            try:
                if not 0 <= i < len(document):
                    raise ValueError("page index is outside the document")
                page = document[i]
                try:
                    width_points = float(page.get_width())
                    height_points = float(page.get_height())
                finally:
                    page.close()
            finally:
                document.close()
        minimum = min(width_points, height_points)
        if not math.isfinite(minimum) or minimum <= 0:
            raise ValueError("PDF page geometry is invalid")
        dpi = max(
            (float(chandra_settings.MIN_PDF_IMAGE_DIM) / minimum) * 72.0,
            float(chandra_settings.IMAGE_DPI),
        )
        expected_width = math.ceil(width_points * dpi / 72.0)
        expected_height = math.ceil(height_points * dpi / 72.0)
    else:
        from PIL import Image

        image_input = io.BytesIO(snapshot_bytes) if snapshot_bytes is not None else path
        with Image.open(image_input) as source:
            source.seek(i)
            expected_width, expected_height = map(int, source.size)
        minimum = min(expected_width, expected_height)
        if minimum <= 0:
            raise ValueError("image geometry is invalid")
        if minimum < int(chandra_settings.MIN_IMAGE_DIM):
            scale = float(chandra_settings.MIN_IMAGE_DIM) / minimum
            expected_width = math.ceil(expected_width * scale)
            expected_height = math.ceil(expected_height * scale)
    expected_pixels = expected_width * expected_height
    if (
        expected_width <= 0
        or expected_height <= 0
        or max(expected_width, expected_height) > MAX_IMAGE_DIMENSION
        or expected_pixels > MAX_IMAGE_PIXELS
        or expected_pixels * 3 * OCR_CONCURRENCY > MAX_CONCURRENT_DECODED_BYTES
    ):
        raise ValueError("page geometry exceeds the pre-render safety limit")
    if snapshot_bytes is None:
        with _PDFIUM_LOCK:
            image = load_file(str(path), {"page_range": str(i)})[0]
    elif path.suffix.lower() == ".pdf":
        with _PDFIUM_LOCK:
            document = pdfium.PdfDocument(snapshot_bytes)
            try:
                document.init_forms()
                page = document[i]
                try:
                    flatten(page)
                    image = page.render(scale=dpi / 72.0).to_pil().convert("RGB")
                    image.load()
                finally:
                    page.close()
            finally:
                document.close()
    else:
        from PIL import Image

        with Image.open(io.BytesIO(snapshot_bytes)) as source:
            source.seek(i)
            image = source.convert("RGB")
            if min(image.width, image.height) < int(chandra_settings.MIN_IMAGE_DIM):
                scale = float(chandra_settings.MIN_IMAGE_DIM) / min(
                    image.width, image.height
                )
                resized = image.resize(
                    (math.ceil(image.width * scale), math.ceil(image.height * scale)),
                    Image.Resampling.LANCZOS,
                )
                image.close()
                image = resized
            image.load()
    pixels = int(image.width) * int(image.height)
    if (
        pixels <= 0
        or pixels > MAX_IMAGE_PIXELS
        or max(int(image.width), int(image.height)) > MAX_IMAGE_DIMENSION
        or pixels * 3 * OCR_CONCURRENCY > MAX_CONCURRENT_DECODED_BYTES
    ):
        close = getattr(image, "close", None)
        if callable(close):
            close()
        raise ValueError(
            f"page {i + 1} expands to {pixels} pixels; limit is {MAX_IMAGE_PIXELS}"
        )
    return image


def _is_reparse_or_symlink(path: Path) -> bool:
    try:
        info = path.lstat()
    except OSError:
        return True
    if stat.S_ISLNK(info.st_mode):
        return True
    attributes = getattr(info, "st_file_attributes", 0)
    reparse = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0)
    return bool(reparse and attributes & reparse)


def _magic_type(path: Path) -> str | None:
    descriptor, before = _open_regular_readonly(path)
    try:
        head = os.read(descriptor, 16)
        after = os.fstat(descriptor)
        current = path.lstat()
        if (
            _stat_signature(before) != _stat_signature(after)
            or not os.path.samestat(after, current)
        ):
            raise RuntimeError("file identity changed during type validation")
    finally:
        os.close(descriptor)
    if head.startswith(b"%PDF-"):
        return "pdf"
    if head.startswith(b"\x89PNG\r\n\x1a\n"):
        return "png"
    if head.startswith(b"\xff\xd8\xff"):
        return "jpeg"
    if head.startswith((b"II*\x00", b"MM\x00*")):
        return "tiff"
    if head.startswith(b"BM"):
        return "bmp"
    return None


def validate_source_file(path: Path, inbox: Path) -> os.stat_result:
    """Perform cheap path checks before making the private immutable snapshot."""
    root = revalidate_runtime_directory(inbox)
    literal = Path(os.path.abspath(os.fspath(path)))
    if literal.parent != root:
        raise ValueError("input must be directly inside this profile's inbox")
    grm_ocr._reject_reparse_chain(literal.parent, "inbox source")
    try:
        info = literal.lstat()
    except OSError as exc:
        raise ValueError("input is missing or inaccessible") from exc
    if (
        not stat.S_ISREG(info.st_mode)
        or grm_ocr._is_link_or_reparse_info(info)
    ):
        raise ValueError("input must be a regular file, not a link or reparse point")
    size = info.st_size
    if size <= 0 or size > MAX_UPLOAD_BYTES:
        raise ValueError(f"input size must be between 1 and {MAX_UPLOAD_BYTES} bytes")
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED:
        raise ValueError(f"unsupported type: {path.name}")
    return info


def _stat_signature(info) -> tuple[int, int, int, int]:
    return (
        int(info.st_dev),
        int(info.st_ino),
        int(info.st_size),
        int(getattr(info, "st_mtime_ns", int(info.st_mtime * 1_000_000_000))),
    )


def _open_regular_readonly(
    path: Path, expected: os.stat_result | None = None
) -> tuple[int, os.stat_result]:
    grm_ocr._reject_reparse_chain(path.parent, "regular file")
    descriptor = grm_ocr._open_no_follow(path)
    try:
        opened = os.fstat(descriptor)
        current = path.lstat()
        if (
            not stat.S_ISREG(opened.st_mode)
            or grm_ocr._is_link_or_reparse_info(opened)
            or grm_ocr._is_link_or_reparse_info(current)
            or (expected is not None and not os.path.samestat(expected, opened))
            or not os.path.samestat(opened, current)
        ):
            raise ValueError("input changed identity or is not a regular file")
        return descriptor, opened
    except Exception:
        os.close(descriptor)
        raise


def _hash_open_file(descriptor: int, maximum: int) -> tuple[str, int]:
    digest = hashlib.sha256()
    total = 0
    while True:
        chunk = os.read(descriptor, 1024 * 1024)
        if not chunk:
            break
        total += len(chunk)
        if total > maximum:
            raise ValueError(f"input grew beyond the {maximum}-byte limit")
        digest.update(chunk)
    return digest.hexdigest(), total


def _expected_magic(suffix: str) -> str:
    return {
        ".pdf": "pdf", ".png": "png", ".jpg": "jpeg", ".jpeg": "jpeg",
        ".tiff": "tiff", ".bmp": "bmp",
    }[suffix]


def create_source_snapshot(
    source: Path, inbox: Path, job_dir: Path
) -> tuple[Path, tuple[int, int, int, int], str, int, bytes]:
    """Copy one stable bounded source generation before invoking native parsers."""
    approved = validate_source_file(source, inbox)
    revalidate_runtime_directory(job_dir.parent)
    grm_ocr._reject_reparse_chain(job_dir, "job directory")
    if not job_dir.is_dir() or _is_reparse_or_symlink(job_dir):
        raise ValueError("job directory is unsafe")
    snapshot = job_dir / f".source{source.suffix.lower()}"
    source_fd = None
    snapshot_fd = None
    try:
        source_fd, before = _open_regular_readonly(source, approved)
        signature = _stat_signature(before)
        if before.st_size <= 0 or before.st_size > MAX_UPLOAD_BYTES:
            raise ValueError(
                f"input size must be between 1 and {MAX_UPLOAD_BYTES} bytes"
            )
        snapshot_fd = os.open(
            snapshot,
            os.O_WRONLY | os.O_CREAT | os.O_EXCL | getattr(os, "O_BINARY", 0),
            0o600,
        )
        digest = hashlib.sha256()
        total = 0
        snapshot_chunks = []
        while True:
            chunk = os.read(source_fd, 1024 * 1024)
            if not chunk:
                break
            total += len(chunk)
            if total > MAX_UPLOAD_BYTES:
                raise ValueError(
                    f"input grew beyond the {MAX_UPLOAD_BYTES}-byte limit while snapshotting"
                )
            digest.update(chunk)
            snapshot_chunks.append(chunk)
            view = memoryview(chunk)
            while view:
                written = os.write(snapshot_fd, view)
                if written <= 0:
                    raise OSError("snapshot write made no progress")
                view = view[written:]
        os.fsync(snapshot_fd)
        after = os.fstat(source_fd)
        current = source.lstat()
        revalidate_runtime_directory(inbox)
        if (
            total != before.st_size
            or _stat_signature(after) != signature
            or not os.path.samestat(after, current)
            or _is_reparse_or_symlink(source)
        ):
            raise RuntimeError("source changed while its immutable snapshot was created")
    except Exception:
        if snapshot_fd is not None:
            os.close(snapshot_fd)
            snapshot_fd = None
        if source_fd is not None:
            os.close(source_fd)
            source_fd = None
        snapshot.unlink(missing_ok=True)
        raise
    finally:
        if snapshot_fd is not None:
            os.close(snapshot_fd)
        if source_fd is not None:
            os.close(source_fd)

    try:
        snapshot_bytes = b"".join(snapshot_chunks)
        if (
            len(snapshot_bytes) != total
            or not hmac.compare_digest(
                hashlib.sha256(snapshot_bytes).hexdigest(), digest.hexdigest()
            )
        ):
            raise RuntimeError("private OCR snapshot did not remain internally consistent")
        detected = _magic_type(snapshot)
        expected = _expected_magic(source.suffix.lower())
        if detected != expected:
            raise ValueError(
                f"file content does not match the {source.suffix.lower()} extension"
            )
        pages = count_pages(snapshot, snapshot_bytes)
        return snapshot, signature, digest.hexdigest(), pages, snapshot_bytes
    except Exception:
        snapshot.unlink(missing_ok=True)
        raise


def verify_source_matches_snapshot(
    source: Path,
    inbox: Path,
    expected_signature: tuple[int, int, int, int],
    expected_digest: str,
) -> None:
    """Refuse completion if the live inbox name no longer names the snapshotted bytes."""
    approved = validate_source_file(source, inbox)
    descriptor = None
    try:
        descriptor, before = _open_regular_readonly(source, approved)
        if _stat_signature(before) != expected_signature:
            raise RuntimeError("source identity or metadata changed during OCR")
        digest, total = _hash_open_file(descriptor, MAX_UPLOAD_BYTES)
        after = os.fstat(descriptor)
        current = source.lstat()
        revalidate_runtime_directory(inbox)
        if (
            total != expected_signature[2]
            or _stat_signature(after) != expected_signature
            or not os.path.samestat(after, current)
            or _is_reparse_or_symlink(source)
            or not hmac.compare_digest(digest, expected_digest)
        ):
            raise RuntimeError("source contents changed during OCR")
    finally:
        if descriptor is not None:
            os.close(descriptor)


def verify_completed_source(path: Path, expected_size: int, expected_digest: str) -> None:
    """Verify the object moved to completion is still the snapshotted generation."""
    descriptor = None
    try:
        descriptor, before = _open_regular_readonly(path)
        digest, total = _hash_open_file(descriptor, MAX_UPLOAD_BYTES)
        after = os.fstat(descriptor)
        if (
            total != expected_size
            or _stat_signature(before) != _stat_signature(after)
            or not hmac.compare_digest(digest, expected_digest)
        ):
            raise RuntimeError("completed source does not match the OCR snapshot")
    finally:
        if descriptor is not None:
            os.close(descriptor)


def _owned_job_tree_size(path: Path) -> int | None:
    """Return a job tree's size, or None if any unsafe node is present."""
    try:
        grm_ocr._reject_reparse_chain(path, "job tree")
    except ValueError:
        return None
    if _is_reparse_or_symlink(path):
        return None
    total = 0
    pending = [path]
    try:
        while pending:
            directory = pending.pop()
            with os.scandir(directory) as entries:
                for entry in entries:
                    child = Path(entry.path)
                    if _is_reparse_or_symlink(child):
                        return None
                    if entry.is_dir(follow_symlinks=False):
                        pending.append(child)
                    elif entry.is_file(follow_symlinks=False):
                        total += entry.stat(follow_symlinks=False).st_size
                    else:
                        return None
        return total
    except OSError:
        return None


def enforce_retention(active_job_id: str | None = None) -> set[str]:
    """Remove expired/oldest job caches until age, count, and byte caps hold."""
    jobs_root = revalidate_runtime_directory(JOBS_DIR)
    now = time.time()
    cutoff = now - RETENTION_DAYS * 86400
    jobs = []
    for path in jobs_root.iterdir():
        try:
            if (
                not JOB_ID_PATTERN.fullmatch(path.name)
                or path.name == active_job_id
                or _is_reparse_or_symlink(path)
                or not path.is_dir()
            ):
                continue
            size = _owned_job_tree_size(path)
            if size is None:
                log(f"retention skipped unsafe job tree: {path.name}")
                continue
            info = path.lstat()
            jobs.append((info.st_mtime, path, size, info))
        except OSError:
            continue
    jobs.sort(reverse=True)
    retained_bytes = sum(size for _, _, size, _ in jobs)
    removed = set()
    for index, (modified, path, size, approved) in enumerate(jobs):
        too_old = modified < cutoff
        over_count = index >= MAX_RETAINED_JOBS
        over_bytes = retained_bytes > MAX_RETAINED_JOB_BYTES
        if not (too_old or over_count or over_bytes):
            continue
        if (
            path.parent != jobs_root
            or not JOB_ID_PATTERN.fullmatch(path.name)
            or _owned_job_tree_size(path) is None
            or not os.path.samestat(approved, path.lstat())
        ):
            log(f"retention refused changed or unsafe job tree: {path.name}")
            continue
        revalidate_runtime_directory(jobs_root)
        grm_ocr._reject_reparse_chain(path, "job tree")
        shutil.rmtree(path)
        revalidate_runtime_directory(jobs_root)
        removed.add(path.name)
        retained_bytes -= size
    if removed:
        with LOCK:
            STATE["history"] = [
                entry for entry in STATE["history"] if entry.get("id") not in removed
            ]
            save_history()
    return removed


def atomic_copy_to_unique(source: Path, directory: Path, name: str) -> Path:
    directory = revalidate_runtime_directory(directory)
    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb", dir=directory, prefix=".copy-", suffix=".tmp", delete=False
        ) as handle:
            temp_path = Path(handle.name)
            descriptor, before = _open_regular_readonly(source)
            try:
                while True:
                    chunk = os.read(descriptor, 1024 * 1024)
                    if not chunk:
                        break
                    handle.write(chunk)
                after = os.fstat(descriptor)
                if _stat_signature(before) != _stat_signature(after):
                    raise RuntimeError("output source changed while copying")
            finally:
                os.close(descriptor)
            handle.flush()
            os.fsync(handle.fileno())
        destination = _publish_staged_unique(temp_path, directory, name)
        temp_path = None
        return destination
    finally:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)


def _candidate_path(directory: Path, name: str, index: int) -> Path:
    base = directory / name
    if index == 0:
        return base
    return directory / f"{base.stem} ({index}){base.suffix}"


def _publish_staged_unique(staged: Path, directory: Path, name: str) -> Path:
    """Atomically publish a same-volume staged file without a visible reservation."""

    directory = revalidate_runtime_directory(directory)
    staged = Path(os.path.abspath(os.fspath(staged)))
    if staged.parent != directory:
        raise RuntimeError("staged publication must stay in its owned destination directory")
    descriptor, opened = _open_regular_readonly(staged)
    destination = None
    try:
        for index in range(1000):
            candidate = _candidate_path(directory, name, index)
            try:
                os.link(staged, candidate, follow_symlinks=False)
            except FileExistsError:
                continue
            destination = candidate
            try:
                linked = candidate.lstat()
                current = staged.lstat()
                revalidate_runtime_directory(directory)
                if (
                    grm_ocr._is_link_or_reparse_info(linked)
                    or not os.path.samestat(opened, linked)
                    or not os.path.samestat(opened, current)
                ):
                    raise RuntimeError("atomic publication changed identity")
                staged.unlink()
                if staged.exists():
                    raise RuntimeError("staged publication could not be finalized")
                final = candidate.lstat()
                if not os.path.samestat(opened, final):
                    raise RuntimeError("published file changed identity")
                return candidate
            except Exception:
                try:
                    current_destination = candidate.lstat()
                    if os.path.samestat(opened, current_destination):
                        candidate.unlink()
                except OSError:
                    pass
                raise
        raise FileExistsError("no collision-free destination name is available")
    finally:
        os.close(descriptor)


def move_source_confirmed(source: Path, destination_dir: Path) -> Path:
    """Link-commit then remove one exact source, so a failed commit never strands it."""

    source = Path(os.path.abspath(os.fspath(source)))
    revalidate_runtime_directory(source.parent)
    destination_dir = revalidate_runtime_directory(destination_dir)
    descriptor, opened = _open_regular_readonly(source)
    try:
        for index in range(1000):
            destination = _candidate_path(destination_dir, source.name, index)
            try:
                os.link(source, destination, follow_symlinks=False)
            except FileExistsError:
                continue
            try:
                linked = destination.lstat()
                current = source.lstat()
                revalidate_runtime_directory(source.parent)
                revalidate_runtime_directory(destination_dir)
                if (
                    grm_ocr._is_link_or_reparse_info(linked)
                    or not os.path.samestat(opened, linked)
                    or not os.path.samestat(opened, current)
                ):
                    raise RuntimeError("source disposition changed identity")
                source.unlink()
                if source.exists() or not os.path.samestat(opened, destination.lstat()):
                    raise RuntimeError("source disposition could not be verified")
                return destination
            except Exception:
                try:
                    current_destination = destination.lstat()
                    if os.path.samestat(opened, current_destination):
                        destination.unlink()
                except OSError:
                    pass
                raise
        raise FileExistsError("no collision-free source destination is available")
    finally:
        os.close(descriptor)


def handle_failed_source(
    source: Path, inbox: Path, error: Exception | str, *, permanent: bool = False
) -> str:
    key = str(source)
    attempts = _retry_counts.get(key, 0) + 1
    _retry_counts[key] = attempts
    if not permanent and attempts < MAX_FILE_RETRIES:
        delay = min(RETRY_DELAY_SECONDS * (2 ** (attempts - 1)), 3600)
        _retry_after[key] = time.time() + delay
        return "retrying"
    if not source.exists():
        _retry_after.pop(key, None)
        _retry_counts.pop(key, None)
        return "missing"
    moved = move_source_confirmed(source, QUARANTINE_DIR)
    reason_path = moved.with_name(moved.name + ".error.json")
    atomic_write_bytes(
        reason_path,
        json.dumps(
            {
                "profile": PROFILE_ID,
                "when": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                "attempts": attempts,
                "error": classified_error(error),
            },
            indent=1,
        ).encode("utf-8"),
    )
    _retry_after.pop(key, None)
    _retry_counts.pop(key, None)
    return "quarantined"


def process_file(src: Path, inbox: Path) -> None:
    revalidate_runtime_directory(inbox)
    processed = revalidate_runtime_directory(PROCESSED_DIR)
    jobs_root = revalidate_runtime_directory(JOBS_DIR)
    job_id = time.strftime("%Y%m%d-%H%M%S") + "-" + secrets.token_hex(4)
    job_dir = jobs_root / job_id
    job_dir.mkdir(parents=True, exist_ok=False, mode=0o700)
    revalidate_runtime_directory(jobs_root)
    grm_ocr._reject_reparse_chain(job_dir, "job directory")
    job_info = job_dir.lstat()
    if not stat.S_ISDIR(job_info.st_mode) or grm_ocr._is_link_or_reparse_info(job_info):
        raise RuntimeError("job directory publication was not a regular directory")
    try:
        job_dir.chmod(0o700)
    except OSError:
        pass

    job = {
        "id": job_id,
        "name": src.name,
        "current_file": src.name,
        "state": "loading",
        "total": 0,
        "done": 0,
        "current": 0,
        "started": time.time(),
        "pages": [],
        "partial": "",
        "regions": [],
        "region_page": 0,
        "error": None,
    }
    with LOCK:
        STATE["job"] = job

    started = time.time()
    page_htmls = []
    page_mds = []
    snapshot_path = None
    snapshot_signature = None
    snapshot_digest = None
    snapshot_bytes = None
    try:
        revalidate_runtime_identity()
        try:
            validate_source_file(src, inbox)
        except ValueError as exc:
            disposition = handle_failed_source(src, inbox, exc, permanent=True)
            raise RuntimeError(f"input rejected and {disposition}") from exc
        try:
            (
                snapshot_path,
                snapshot_signature,
                snapshot_digest,
                n_pages,
                snapshot_bytes,
            ) = create_source_snapshot(src, inbox, job_dir)
        except ValueError as exc:
            disposition = handle_failed_source(src, inbox, exc, permanent=True)
            raise RuntimeError(f"input rejected and {disposition}") from exc
        # preflight: a dead OCR server should fail the job in seconds with a
        # clear message, not grind a per-page timeout for every page
        try:
            grm_ocr.probe(timeout=8)
        except Exception as exc:
            raise RuntimeError(
                "OCR server is unavailable or its TLS/configuration check failed. "
                "The document was left in the inbox and will retry when the server is back."
            ) from exc

        job["total"] = n_pages
        job["pages"] = [
            {"n": i + 1, "file": src.name, "state": "pending", "secs": None, "chars": None}
            for i in range(n_pages)
        ]
        job["state"] = "ocr"
        page_mds = [""] * n_pages
        page_htmls = [""] * n_pages
        if n_pages > MAX_REMOTE_ATTEMPTS:
            raise ValueError("document exceeds the aggregate remote request budget")
        attempts_per_page = min(2, max(1, MAX_REMOTE_ATTEMPTS // n_pages))
        output_budget_lock = threading.Lock()
        output_chars = [0]

        def refresh_current():
            """Live pane follows the earliest page still being read."""
            for p in job["pages"]:
                if p["state"] == "working":
                    if job["current"] != p["n"]:
                        job["current"] = p["n"]
                        job["partial"] = ""
                        job["regions"] = []
                        job["region_page"] = p["n"]
                    return
            job["current"] = job["done"]

        def do_page(i):
            page = job["pages"][i]
            if job.get("cancel"):
                page["state"] = "skipped"
                with LOCK:
                    job["done"] += 1
                return
            try:
                img = load_page(snapshot_path, i, snapshot_bytes)
            except Exception as e:
                page["state"] = "error"
                page["error"] = classified_error(e)
                with LOCK:
                    job["done"] += 1
                    refresh_current()
                return
            # display JPEG just-in-time — the first beam appears in seconds,
            # not after a long doc has fully pre-rendered
            display_image = None
            encoded_image = None
            try:
                disp = img
                if img.width > DISPLAY_WIDTH:
                    display_image = img.resize(
                        (DISPLAY_WIDTH, int(img.height * DISPLAY_WIDTH / img.width))
                    )
                    disp = display_image
                encoded_image = disp.convert("RGB")
                encoded_image.save(job_dir / f"page_{i + 1}.jpg", quality=82)
            except Exception as e:
                log(f"page {i + 1} render failed: {classified_error(e)}")
            finally:
                if encoded_image is not None and encoded_image is not img:
                    close = getattr(encoded_image, "close", None)
                    if callable(close):
                        close()
                if display_image is not None and display_image is not encoded_image:
                    close = getattr(display_image, "close", None)
                    if callable(close):
                        close()
            page["state"] = "working"
            with LOCK:
                refresh_current()
            t0 = time.time()
            last = [0.0]

            def on_delta(raw_so_far):
                if job.get("cancel"):
                    raise _Cancelled()
                if len(raw_so_far) > MAX_PAGE_OUTPUT_CHARS:
                    raise ValueError("OCR page output exceeded its safety limit")
                if time.time() - t0 > MAX_PAGE_OCR_SECONDS * attempts_per_page:
                    raise TimeoutError("OCR page exceeded its wall-clock limit")
                # only the followed page pays the parse cost
                if job["current"] != i + 1:
                    return
                now = time.time()
                if now - last[0] < 0.4:
                    return
                last[0] = now
                try:
                    md = grm_ocr.raw_to_markdown(raw_so_far)
                except Exception:
                    md = ""
                if not md or md.lstrip().startswith(("<!DOCTYPE", "<html", "<script")):
                    md = BeautifulSoup(
                        grm_ocr.normalize_raw(raw_so_far), "html.parser"
                    ).get_text("\n")
                if len(md) > MAX_PAGE_OUTPUT_CHARS:
                    raise ValueError("OCR page preview exceeded its safety limit")
                if job["current"] == i + 1:
                    regions = extract_regions(raw_so_far)
                    with LOCK:
                        if job["current"] == i + 1:
                            job["partial"] = md
                            job["regions"] = regions
                            job["region_page"] = i + 1

            try:
                raw = grm_ocr.ocr_page_raw(
                    img,
                    on_delta=on_delta,
                    max_retries=attempts_per_page - 1,
                )
                if len(raw) > MAX_PAGE_OUTPUT_CHARS:
                    raise ValueError("OCR page output exceeded its safety limit")
                md = grm_ocr.raw_to_markdown(raw)
                html = sanitize_ocr_html(grm_ocr.raw_to_html(raw))
                page_chars = len(md) + len(html)
                if len(md) > MAX_PAGE_OUTPUT_CHARS or len(html) > MAX_PAGE_OUTPUT_CHARS:
                    raise ValueError("OCR page conversion exceeded its safety limit")
                with output_budget_lock:
                    if output_chars[0] + page_chars > MAX_JOB_OUTPUT_CHARS:
                        job["cancel"] = True
                        raise ValueError("OCR document output exceeded its safety limit")
                    output_chars[0] += page_chars
                (job_dir / f"page_{i + 1}.md").write_text(md, encoding="utf-8")
                (job_dir / f"page_{i + 1}.html").write_text(html, encoding="utf-8")
                page_mds[i] = md
                page_htmls[i] = html
                page["state"] = "done"
                page["secs"] = round(time.time() - t0, 1)
                page["chars"] = len(md)
                regions = extract_regions(raw)
                with LOCK:
                    if job["current"] == i + 1:
                        job["regions"] = regions
                        job["region_page"] = i + 1
            except _Cancelled:
                page["state"] = "skipped"
                page["secs"] = round(time.time() - t0, 1)
            except Exception as e:
                page["state"] = "error"
                page["secs"] = round(time.time() - t0, 1)
                page["error"] = classified_error(e)
            finally:
                close = getattr(img, "close", None)
                if callable(close):
                    close()
            with LOCK:
                job["done"] += 1
                refresh_current()

        with ThreadPoolExecutor(max_workers=OCR_CONCURRENCY) as pool:
            list(pool.map(do_page, range(n_pages)))
        revalidate_runtime_identity()
        verify_source_matches_snapshot(
            src, inbox, snapshot_signature, snapshot_digest
        )
        job["partial"] = ""

        cancelled = bool(job.get("cancel"))
        errors = sum(1 for p in job["pages"] if p["state"] not in ("done", "skipped"))
        # Build job artifacts first. Processed receives atomic copies only after
        # every export has closed successfully.
        stem = Path(sanitize_name(src.name)).stem
        job_txt = job_dir / f"{stem}.txt"
        atomic_write_bytes(job_txt, export_txt(page_htmls).encode("utf-8"))
        job_xlsx = job_dir / f"{stem}.xlsx"
        try:
            export_xlsx(page_htmls, job_xlsx)
        except Exception as e:
            log(f"xlsx export failed for {src.name}: {classified_error(e)}")
            job_xlsx = None

        txt_out = None
        xlsx_out = None
        committed = []
        dest_dir = processed
        if cancelled:
            dest_dir = ON_HOLD_DIR
        elif errors:
            dest_dir = NEEDS_REVIEW_DIR
        try:
            verify_source_matches_snapshot(
                src, inbox, snapshot_signature, snapshot_digest
            )
            txt_out = atomic_copy_to_unique(job_txt, processed, job_txt.name)
            committed.append(txt_out)
            if job_xlsx:
                xlsx_out = atomic_copy_to_unique(job_xlsx, processed, job_xlsx.name)
                committed.append(xlsx_out)
            revalidate_runtime_identity()
            verify_source_matches_snapshot(
                src, inbox, snapshot_signature, snapshot_digest
            )
            dest = move_source_confirmed(src, dest_dir)
            try:
                verify_completed_source(
                    dest, snapshot_signature[2], snapshot_digest
                )
            except Exception as verify_error:
                quarantined = move_source_confirmed(dest, QUARANTINE_DIR)
                atomic_write_bytes(
                    quarantined.with_name(quarantined.name + ".error.json"),
                    json.dumps(
                        {
                            "profile": PROFILE_ID,
                            "when": time.strftime(
                                "%Y-%m-%dT%H:%M:%SZ", time.gmtime()
                            ),
                            "attempts": 1,
                            "error": "completed source changed after OCR snapshot",
                        },
                        indent=1,
                    ).encode("utf-8"),
                )
                raise RuntimeError(
                    "completed source did not match the OCR snapshot and was quarantined"
                ) from verify_error
            revalidate_runtime_identity()
        except Exception:
            for path in committed:
                try:
                    path.unlink(missing_ok=True)
                except OSError:
                    log(f"rollback could not remove incomplete output: {path.name}")
            raise

        if cancelled:
            job["state"] = "cancelled"
        elif errors:
            job["state"] = "finished_with_errors"
        else:
            job["state"] = "finished"
        _retry_after.pop(str(src), None)
        _retry_counts.pop(str(src), None)
        entry = {
            "id": job_id,
            "name": src.name,
            "status": job["state"],
            "when": time.strftime("%Y-%m-%d %H:%M"),
            "pages": job["total"],
            "errors": errors,
            "secs": round(time.time() - started, 0),
            "chars": sum(p["chars"] or 0 for p in job["pages"]),
            "links": {
                "md": f"/jobs/{job_id}/{urllib.parse.quote(job_txt.name)}",
                "xlsx": (
                    f"/jobs/{job_id}/{urllib.parse.quote(job_xlsx.name)}"
                    if job_xlsx else None
                ),
            },
        }
        with LOCK:
            STATE["history"].insert(0, entry)
            STATE["history"] = STATE["history"][:MAX_RETAINED_JOBS]
            save_history()
        enforce_retention(active_job_id=job_id)
        log(f"done: {src.name} ({job['total']} pages, {errors} errors)")
    except Exception as e:
        disposition = "missing" if not src.exists() else "failed"
        if src.exists():
            try:
                disposition = handle_failed_source(src, inbox, e)
            except Exception as quarantine_error:
                log(
                    f"failure disposition could not complete for {src.name}: "
                    f"{classified_error(quarantine_error)}"
                )
                disposition = "failed"
        job["state"] = disposition
        job["error"] = classified_error(e)
        log(f"FAILED ({disposition}): {src.name}: {classified_error(e)}")
    finally:
        if snapshot_path is not None:
            try:
                snapshot_path.unlink(missing_ok=True)
            except OSError as cleanup_error:
                log(
                    f"snapshot cleanup failed for {src.name}: "
                    f"{classified_error(cleanup_error)}"
                )
        time.sleep(2)  # let the UI show the finished state before clearing
        with LOCK:
            if STATE["job"] is job:
                STATE["job"] = None


def worker(inbox: Path):
    while True:
        src = None
        with LOCK:
            if _pending_paths:
                src = _pending_paths.pop(0)
                STATE["queue"].pop(0)
        if src is None:
            time.sleep(1)
            continue
        if src.exists():
            process_file(src, inbox)


def prune_watcher_sizes(sizes: dict, present: set[Path]) -> None:
    """Bound and prune watcher stability state after names disappear."""
    for path in list(sizes):
        if path not in present or not path.exists():
            sizes.pop(path, None)
    if len(sizes) > MAX_WATCHER_TRACKED:
        for path in sorted(sizes, key=lambda value: str(value))[MAX_WATCHER_TRACKED:]:
            sizes.pop(path, None)


def watcher(inbox: Path):
    """Enqueue new files once their size is stable (copy finished)."""
    sizes = {}
    while True:
        present = set()
        try:
            owned_inbox = revalidate_runtime_directory(inbox)
            for p in sorted(owned_inbox.iterdir()):
                if (
                    _is_reparse_or_symlink(p)
                    or not p.is_file()
                    or p.suffix.lower() not in SUPPORTED
                ):
                    continue
                present.add(p)
                if _retry_after.get(str(p), 0) > time.time():
                    continue
                with LOCK:
                    queued = any(q == p for q in _pending_paths) or (
                        STATE["job"] and STATE["job"]["name"] == p.name
                        and STATE["job"]["state"] not in ("finished", "failed")
                    )
                if queued:
                    continue
                size = p.stat().st_size
                if sizes.get(p) == size and size > 0:
                    with LOCK:
                        _pending_paths.append(p)
                        STATE["queue"].append({"name": p.name, "size": size})
                    log(f"queued: {p.name}")
                    sizes.pop(p, None)
                else:
                    if p in sizes or len(sizes) < MAX_WATCHER_TRACKED:
                        sizes[p] = size
            revalidate_runtime_directory(owned_inbox)
        except Exception as e:
            log(f"watcher error: {classified_error(e)}")
        finally:
            prune_watcher_sizes(sizes, present)
        time.sleep(2)


# ---------------------------------------------------------------- http server

class Handler(SimpleHTTPRequestHandler):
    protocol_version = "HTTP/1.0"

    def log_message(self, *a):
        pass

    def _json(self, obj, code=200, extra_headers=None):
        body = json.dumps(obj).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        if extra_headers:
            for key, value in extra_headers.items():
                self.send_header(key, value)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _request_host(self):
        values = self.headers.get_all("Host", [])
        if len(values) != 1:
            return None
        rendered = values[0].strip()
        if not rendered or any(ch.isspace() for ch in rendered):
            return None
        try:
            parsed = urllib.parse.urlsplit("//" + rendered)
            host = parsed.hostname.lower().rstrip(".")
            port = parsed.port or self.server.server_port
        except (AttributeError, ValueError):
            return None
        if parsed.username or parsed.path or parsed.query or parsed.fragment:
            return None
        return host, port

    def _origin_allowed(self, origin: str) -> bool:
        if origin in getattr(self.server, "allowed_origins", set()):
            return True
        if origin == "null":
            return False
        parsed = urllib.parse.urlsplit(origin)
        if parsed.scheme not in {"http", "https"} or not parsed.hostname:
            return False
        request = self._request_host()
        if request is None:
            return False
        request_host, request_port = request
        origin_port = parsed.port or (443 if parsed.scheme == "https" else 80)
        return (
            parsed.username is None
            and not parsed.path.rstrip("/")
            and not parsed.query
            and not parsed.fragment
            and parsed.hostname.lower().rstrip(".") == request_host
            and origin_port == request_port
        )

    def end_headers(self):
        self.send_header("Cache-Control", "private, no-store, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "no-referrer")
        self.send_header("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
        self.send_header(
            "Content-Security-Policy",
            "default-src 'self'; img-src 'self' data: blob:; "
            "style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'; "
            "connect-src 'self'; object-src 'none'; frame-ancestors 'none'; "
            "base-uri 'none'; form-action 'none'",
        )
        origin = self.headers.get("Origin")
        if origin and self._origin_allowed(origin):
            self.send_header("Access-Control-Allow-Origin", origin)
            self.send_header("Vary", "Origin")
        super().end_headers()

    def _authorized(self) -> bool:
        expected = self.server.auth_token
        authorizations = self.headers.get_all("Authorization", [])
        header_tokens = self.headers.get_all("X-Document-Reader-Token", [])
        if len(authorizations) > 1 or len(header_tokens) > 1:
            return False
        authorization = authorizations[0] if authorizations else ""
        bearer = ""
        basic_user = ""
        basic_token = ""
        scheme, separator, credentials = authorization.partition(" ")
        if separator and scheme.lower() == "bearer":
            bearer = credentials.strip()
        elif separator and scheme.lower() == "basic":
            try:
                decoded = base64.b64decode(credentials.strip(), validate=True).decode("utf-8")
                basic_user, basic_token = decoded.split(":", 1)
            except (ValueError, UnicodeError):
                pass
        header_token = header_tokens[0] if header_tokens else ""
        bearer_ok = hmac.compare_digest(expected, bearer)
        header_ok = hmac.compare_digest(expected, header_token)
        basic_ok = hmac.compare_digest(expected, basic_token) & hmac.compare_digest(
            self.server.profile_id, basic_user
        )
        return bool(bearer_ok | header_ok | basic_ok)

    def _owner_header_authorized(self) -> bool:
        expected = self.server.owner_fingerprint
        supplied = self.headers.get_all("X-Document-Reader-Owner", [])
        return len(supplied) == 1 and hmac.compare_digest(expected, supplied[0])

    def _owner_authorized(self) -> bool:
        supplied = self.headers.get_all("X-Document-Reader-Owner", [])
        if supplied:
            return self._owner_header_authorized()
        expected = self.server.owner_fingerprint
        cookie_headers = self.headers.get_all("Cookie", [])
        if len(cookie_headers) != 1 or len(cookie_headers[0]) > MAX_HEADER_VALUE:
            return False
        try:
            cookies = SimpleCookie()
            cookies.load(cookie_headers[0])
        except CookieError:
            return False
        morsel = cookies.get(self.server.owner_cookie_name)
        return bool(morsel and hmac.compare_digest(expected, morsel.value))

    def _request_policy(self, *, require_auth=True) -> bool:
        try:
            revalidate_runtime_identity(getattr(self.server, "runtime_identity", None))
        except RuntimeError:
            self._json({"error": "profile runtime identity changed"}, 409)
            return False
        if len(self.path) > MAX_REQUEST_PATH:
            self._json({"error": "request target too long"}, 414)
            return False
        try:
            self._parsed_request_target = _strict_request_target(self.path)
        except ValueError:
            self._json({"error": "invalid request target"}, 400)
            return False
        pairs = list(self.headers.items())
        if (
            len(pairs) > MAX_HEADER_COUNT
            or sum(len(key) + len(value) for key, value in pairs) > MAX_HEADER_BYTES
            or any(len(value) > MAX_HEADER_VALUE for _, value in pairs)
        ):
            self._json({"error": "request headers too large"}, 431)
            return False
        host = self._request_host()
        if host is None or host[0] not in self.server.allowed_hosts:
            self._json({"error": "unrecognized Host"}, 421)
            return False
        origins = self.headers.get_all("Origin", [])
        if len(origins) > 1 or (origins and not self._origin_allowed(origins[0])):
            self._json({"error": "origin rejected"}, 403)
            return False
        if require_auth and not self._authorized():
            self._json(
                {"error": "authentication required"},
                401,
                {"WWW-Authenticate": f'Basic realm="{SERVICE_NAME}:{self.server.profile_id}"'},
            )
            return False
        route = self._parsed_request_target.path
        if route in {"/api/health", "/api/shutdown"}:
            if not self._owner_header_authorized():
                self._json({"error": "owner identity required"}, 403)
                return False
        elif route.startswith("/api/") or route.startswith("/jobs/"):
            if not self._owner_authorized():
                self._json({"error": "owner identity required"}, 403)
                return False
        return True

    def do_OPTIONS(self):
        if not self._request_policy(require_auth=False):
            return
        self.send_response(204)
        self.send_header("Allow", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header(
            "Access-Control-Allow-Headers",
            "Authorization, Content-Type, X-Document-Reader-Token, "
            "X-Document-Reader-Owner",
        )
        self.send_header("Access-Control-Max-Age", "300")
        self.send_header("Content-Length", "0")
        self.end_headers()

    def do_GET(self):
        if not self._request_policy():
            return
        parsed = self._parsed_request_target
        path = parsed.path
        if parsed.query or parsed.fragment:
            return self._json({"error": "unexpected query"}, 400)
        if path == "/api/health":
            return self._json(
                {
                    "status": "ok",
                    "service": SERVICE_NAME,
                    "version": VERSION,
                    "api_version": API_VERSION,
                    "profile_name": self.server.profile_id,
                    "owner_fingerprint": self.server.owner_fingerprint,
                    "instance_id": self.server.instance_id,
                    "port": self.server.server_port,
                    "pid": os.getpid(),
                    "started_at": self.server.started_at,
                }
            )
        if path == "/api/state":
            with LOCK:
                # deep copy under the lock — the worker threads mutate the
                # job dict continuously; a shallow dict() hands the JSON
                # encoder live-mutating page lists
                snap = copy.deepcopy(
                    {
                        "version": VERSION,
                        "profile": self.server.profile_id,
                        "queue": STATE["queue"],
                        "job": STATE["job"],
                        "history": STATE["history"][:30],
                    }
                )
            if snap["job"]:
                snap["job"]["base"] = f"/jobs/{snap['job']['id']}"
            return self._json(snap)
        if path in {"/", "/index.html"}:
            return self._file(VIEWER_DIR / "firm.html", "text/html; charset=utf-8")
        if path.startswith("/jobs/"):
            relative = path[len("/jobs/") :]
            parts = relative.split("/")
            if (
                len(parts) != 2
                or not JOB_ID_PATTERN.fullmatch(parts[0])
                or not parts[1]
                or parts[1].startswith(".")
                or Path(parts[1]).name != parts[1]
                or Path(parts[1]).suffix.lower() not in JOB_DOWNLOAD_SUFFIXES
            ):
                return self._json({"error": "bad path"}, 400)
            target = JOBS_DIR / parts[0] / parts[1]
            ctype = {
                ".jpg": "image/jpeg", ".md": "text/markdown; charset=utf-8",
                ".txt": "text/plain; charset=utf-8",
                ".html": "text/html; charset=utf-8",
                ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }.get(target.suffix.lower(), "application/octet-stream")
            return self._file(target, ctype)
        return self._json({"error": "not found"}, 404)

    def _file(self, p: Path, ctype: str):
        p = Path(os.path.abspath(os.fspath(p)))
        viewer_file = p.parent == Path(os.path.abspath(os.fspath(VIEWER_DIR)))
        is_job = False
        if viewer_file:
            if p.name != "firm.html":
                return self._json({"error": "not found"}, 404)
            grm_ocr._reject_reparse_chain(p.parent, "viewer asset")
        else:
            jobs_root = revalidate_runtime_directory(JOBS_DIR)
            if (
                p.parent.parent != jobs_root
                or not JOB_ID_PATTERN.fullmatch(p.parent.name)
            ):
                return self._json({"error": "not found"}, 404)
            grm_ocr._reject_reparse_chain(p.parent, "job asset")
            if _is_reparse_or_symlink(p.parent):
                return self._json({"error": "not found"}, 404)
            is_job = True
        try:
            descriptor, opened = _open_regular_readonly(p)
        except (OSError, ValueError):
            return self._json({"error": "not found"}, 404)
        try:
            if p.suffix.lower() == ".html" and is_job:
                if opened.st_size > MAX_INLINE_HTML_BYTES:
                    return self._json({"error": "result is too large to render"}, 413)
                chunks = []
                remaining = MAX_INLINE_HTML_BYTES + 1
                while True:
                    chunk = os.read(descriptor, min(1024 * 1024, remaining))
                    if not chunk:
                        break
                    chunks.append(chunk)
                    remaining -= len(chunk)
                    if remaining <= 0:
                        return self._json({"error": "result is too large to render"}, 413)
                data = sanitize_ocr_html(
                    b"".join(chunks).decode("utf-8", errors="replace")
                ).encode("utf-8")
                length = len(data)
            else:
                data = None
                length = opened.st_size
            self.send_response(200)
            self.send_header("Content-Type", ctype)
            if viewer_file:
                self.send_header(
                    "Set-Cookie",
                    f"{self.server.owner_cookie_name}={self.server.owner_fingerprint}; "
                    "Path=/; HttpOnly; SameSite=Strict; Max-Age=3600",
                )
            if p.suffix.lower() in {".xlsx", ".md", ".txt"} and is_job:
                fallback = re.sub(r"[^A-Za-z0-9._-]", "_", p.name)[:120] or "download"
                encoded = urllib.parse.quote(p.name, safe="")
                self.send_header(
                    "Content-Disposition", f"attachment; filename=\"{fallback}\"; filename*=UTF-8''{encoded}"
                )
            self.send_header("Content-Length", str(length))
            self.end_headers()
            if data is not None:
                self.wfile.write(data)
                return
            while True:
                chunk = os.read(descriptor, 1024 * 1024)
                if not chunk:
                    break
                revalidate_runtime_identity(self.server.runtime_identity)
                self.wfile.write(chunk)
        except (BrokenPipeError, ConnectionResetError, socket.timeout):
            pass
        finally:
            os.close(descriptor)
        return

    def do_POST(self):
        if not self._request_policy():
            return
        path = self._parsed_request_target
        if path.path == "/api/shutdown":
            if path.query:
                return self._json({"error": "unexpected query"}, 400)
            self._json({"ok": True})
            if self.server.hard_exit_on_shutdown:
                def terminate_process():
                    time.sleep(0.05)
                    os._exit(0)

                threading.Thread(target=terminate_process, daemon=True).start()
            else:
                threading.Thread(target=self.server.shutdown, daemon=True).start()
            return
        if path.path == "/api/cancel":
            if path.query:
                return self._json({"error": "unexpected query"}, 400)
            with LOCK:
                job = STATE["job"]
                terminal = {
                    "finished", "finished_with_errors", "failed", "cancelled", "quarantined"
                }
                if job and job["state"] not in terminal:
                    job["cancel"] = True
                    log(f"cancel requested: {job['name']}")
                    return self._json({"ok": True})
            return self._json({"ok": False, "error": "no active job"}, 409)
        if path.path != "/api/upload":
            return self._json({"error": "not found"}, 404)
        if self.headers.get("Transfer-Encoding"):
            return self._json({"error": "chunked transfer encoding is not accepted"}, 400)
        try:
            query = urllib.parse.parse_qs(path.query, strict_parsing=True, max_num_fields=4)
        except ValueError:
            return self._json({"error": "bad query"}, 400)
        raw_name = query.get("name", ["upload.pdf"])[0]
        if len(raw_name) > 180:
            return self._json({"error": "file name too long"}, 400)
        name = sanitize_name(raw_name)
        suffix = Path(name).suffix.lower()
        if suffix not in SUPPORTED:
            return self._json({"error": f"unsupported type: {name}"}, 400)
        length_values = self.headers.get_all("Content-Length", [])
        if len(length_values) != 1 or not length_values[0].isdigit():
            return self._json({"error": "valid Content-Length required"}, 411)
        length = int(length_values[0])
        maximum = self.server.max_upload_bytes
        if length <= 0 or length > maximum:
            return self._json({"error": f"upload must be between 1 and {maximum} bytes"}, 413)
        # Stream only to an unsupported suffix, then atomically publish a
        # collision-free hard link. The watcher never sees a reservation.
        revalidate_runtime_directory(self.server.inbox)
        dest = None
        temp_path = None
        remaining = length
        committed = False
        rejection = None
        try:
            with tempfile.NamedTemporaryFile(
                mode="wb",
                dir=self.server.inbox,
                prefix=".upload-",
                suffix=".uploading",
                delete=False,
            ) as handle:
                temp_path = Path(handle.name)
                while remaining > 0:
                    revalidate_runtime_identity(self.server.runtime_identity)
                    chunk = self.rfile.read(min(65536, remaining))
                    if not chunk:
                        break
                    handle.write(chunk)
                    remaining -= len(chunk)
                handle.flush()
                os.fsync(handle.fileno())
            if remaining:
                rejection = ({"error": "incomplete upload"}, 400)
            else:
                expected = {
                    ".pdf": "pdf",
                    ".png": "png",
                    ".jpg": "jpeg",
                    ".jpeg": "jpeg",
                    ".tiff": "tiff",
                    ".bmp": "bmp",
                }[suffix]
                if _magic_type(temp_path) != expected:
                    rejection = (
                        {"error": "file content does not match its extension"},
                        415,
                    )
                else:
                    revalidate_runtime_identity(self.server.runtime_identity)
                    dest = _publish_staged_unique(
                        temp_path, self.server.inbox, name
                    )
                    temp_path = None
                    committed = True
        except socket.timeout:
            rejection = ({"error": "upload timed out"}, 408)
        finally:
            if temp_path is not None:
                temp_path.unlink(missing_ok=True)
        if rejection is not None:
            return self._json(*rejection)
        log(f"uploaded: {dest.name}")
        return self._json({"ok": True, "name": dest.name})


class BoundedThreadingHTTPServer(ThreadingHTTPServer):
    daemon_threads = True

    def __init__(self, server_address, handler, *, max_threads=MAX_HTTP_THREADS):
        self._request_slots = threading.BoundedSemaphore(max_threads)
        super().__init__(server_address, handler)

    def get_request(self):
        request, address = super().get_request()
        request.settimeout(REQUEST_SOCKET_TIMEOUT)
        return request, address

    def process_request(self, request, client_address):
        if not self._request_slots.acquire(blocking=False):
            try:
                request.sendall(
                    b"HTTP/1.0 503 Service Unavailable\r\nConnection: close\r\n"
                    b"Content-Length: 0\r\n\r\n"
                )
            except OSError:
                pass
            self.shutdown_request(request)
            return
        try:
            super().process_request(request, client_address)
        except Exception:
            self._request_slots.release()
            raise

    def process_request_thread(self, request, client_address):
        try:
            super().process_request_thread(request, client_address)
        finally:
            self._request_slots.release()


class DualStackHTTPServer(BoundedThreadingHTTPServer):
    """Listen on [::] with IPV6_V6ONLY off so both ::1 and 127.0.0.1 work."""

    address_family = socket.AF_INET6

    def server_bind(self):
        try:
            self.socket.setsockopt(socket.IPPROTO_IPV6, socket.IPV6_V6ONLY, 0)
        except OSError:
            pass
        super().server_bind()


def _normalized_host(value: str) -> str:
    value = str(value).strip()
    if not value or any(ch.isspace() for ch in value):
        raise ValueError(f"invalid allowed host: {value!r}")
    try:
        parsed = urllib.parse.urlsplit("//" + value)
        host = parsed.hostname.lower().rstrip(".")
    except (AttributeError, ValueError) as exc:
        raise ValueError(f"invalid allowed host: {value!r}") from exc
    if parsed.username or parsed.path or parsed.query or parsed.fragment:
        raise ValueError(f"invalid allowed host: {value!r}")
    return host


def build_server(
    bind,
    port,
    *,
    auth_token,
    profile=PROFILE_ID,
    data_root=None,
    owner_fingerprint=None,
    instance_id=None,
    started_at=None,
    runtime_identity=None,
    hard_exit_on_shutdown=False,
    allow_remote=False,
    allowed_hosts=None,
    allowed_origins=None,
    max_threads=MAX_HTTP_THREADS,
):
    profile = normalize_profile(profile)
    if not strong_token(auth_token):
        raise ValueError("a strong document-reader token is required")
    owner_fingerprint = str(owner_fingerprint or "")
    instance_id = str(instance_id or "")
    if not re.fullmatch(r"[0-9a-f]{64}", owner_fingerprint):
        raise ValueError("owner fingerprint must be 64 lowercase hex characters")
    if not re.fullmatch(r"[0-9a-f]{32}", instance_id):
        raise ValueError("instance id must be 32 lowercase hex characters")
    remote = not is_loopback_bind(bind)
    if remote and not allow_remote:
        raise ValueError("non-loopback bind requires --allow-remote")
    hosts = {_normalized_host(value) for value in (allowed_hosts or [])}
    if not remote:
        hosts.update({"localhost", "127.0.0.1", "::1"})
    elif bind not in {"", "0.0.0.0", "::"}:
        hosts.add(_normalized_host(bind))
    if remote and not hosts:
        raise ValueError("non-loopback bind requires at least one --allowed-host")
    origins = {str(value).strip() for value in (allowed_origins or []) if str(value).strip()}
    max_threads = bounded_int(max_threads, "max HTTP threads", 1, 64)

    server_class = BoundedThreadingHTTPServer
    address = (bind, port)
    if bind in {"", "::"}:
        server_class = DualStackHTTPServer
        address = ("::", port)
    server = server_class(address, Handler, max_threads=max_threads)
    server.auth_token = auth_token
    server.profile_id = profile
    server.owner_fingerprint = owner_fingerprint
    server.owner_cookie_name = f"DocumentReaderOwner_{owner_fingerprint[:16]}"
    server.instance_id = instance_id
    server.started_at = started_at or time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    server.runtime_identity = runtime_identity
    server.hard_exit_on_shutdown = bool(hard_exit_on_shutdown)
    server.allowed_hosts = hosts
    server.allowed_origins = origins
    server.max_upload_bytes = MAX_UPLOAD_BYTES
    return server


def run_service_config(config_path: Path) -> None:
    global _RUNTIME_IDENTITY, _RUNTIME_CHILD_IDENTITIES
    config = load_service_config(Path(config_path))
    token = read_service_token(Path(config["token_file"]))
    plugin_root = Path(config["plugin_root"])
    lock_path = plugin_root / "runtime" / "service.lock"
    started_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    with RuntimeOwnerLock(lock_path):
        engine_config = grm_ocr.load_profile_config(
            plugin_root / "config" / "engine.json",
            plugin_root / "config" / "engine.token",
        )
        engine_config = dataclasses.replace(
            engine_config,
            request_timeout=min(engine_config.request_timeout, 60),
            transport_retries=0,
        )
        grm_ocr.configure(engine_config)
        root = configure_runtime(
            config["profile"],
            Path(config["data_root"]),
            jobs_root=Path(config["jobs"]),
            state_root=Path(config["state"]),
            processed_root=Path(config["processed"]),
        )
        inbox = Path(config["inbox"]).resolve()
        for directory in (inbox, Path(config["logs"])):
            if not directory.is_relative_to(root):
                raise ValueError("configured runtime directory escapes the profile data root")
            _private_directory(directory)
        _RUNTIME_IDENTITY = capture_runtime_identity(
            root, config["owner_id"], config["profile"]
        )
        _RUNTIME_CHILD_IDENTITIES = capture_runtime_directories(
            {
                "inbox": inbox,
                "processed": PROCESSED_DIR,
                "jobs": JOBS_DIR,
                "state": STATE_DIR,
                "logs": Path(config["logs"]),
                "on-hold": ON_HOLD_DIR,
                "needs-review": NEEDS_REVIEW_DIR,
                "quarantine": QUARANTINE_DIR,
            }
        )
        load_history()
        enforce_retention()

        server = build_server(
            config["bind"],
            config["port"],
            auth_token=token,
            profile=config["profile"],
            data_root=root,
            owner_fingerprint=config["owner_id"],
            instance_id=config["instance_id"],
            started_at=started_at,
            runtime_identity=_RUNTIME_IDENTITY,
            hard_exit_on_shutdown=True,
            allow_remote=False,
            allowed_hosts=[],
            allowed_origins=[],
            max_threads=MAX_HTTP_THREADS,
        )
        server.inbox = inbox
        threading.Thread(target=watcher, args=(inbox,), daemon=True).start()
        threading.Thread(target=worker, args=(inbox,), daemon=True).start()
        log(
            f"Document Reader {VERSION} profile={config['profile']} on "
            f"http://127.0.0.1:{config['port']} owner={config['owner_id'][:12]}"
        )
        try:
            server.serve_forever()
        finally:
            server.server_close()
            grm_ocr.configure(None)
            _RUNTIME_IDENTITY = None
            _RUNTIME_CHILD_IDENTITIES = {}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", required=True, type=Path)
    args = parser.parse_args()
    run_service_config(args.config)


if __name__ == "__main__":
    main()
